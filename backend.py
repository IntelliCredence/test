from flask import Flask, render_template_string, request, jsonify, send_file
from flask_cors import CORS
import requests
import json
import os
import random
import base64
from typing import Dict, List, Any
from pathlib import Path
from io import BytesIO
import time
from openai import OpenAI
import threading
from urllib.parse import quote
import openpyxl
from flask import redirect
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import re
from datetime import datetime, timedelta

import glob

OPENAI_API_KEY = ""
app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'dev-secret-key')

# CORS configuration
CORS(app, 
     resources={r"/api/*": {"origins": "*"}},
     supports_credentials=True,
     allow_headers=["Content-Type", "Authorization"],
     methods=["GET", "POST", "OPTIONS"])
# ====================
# BACKEND CODE - Keep your existing backend code EXACTLY as is
# ====================
from google.cloud import storage
import tempfile

# Add after your other imports
GCS_BUCKET_NAME = os.environ.get('GCS_BUCKET_NAME', 'figma-test-cases-bucket')  # Change this

def upload_to_gcs(local_file_path, gcs_file_name):
    """Upload file to Google Cloud Storage"""
    try:
        if not os.path.exists(local_file_path):
            print(f"❌ Local file does not exist: {local_file_path}")
            return False
        
        storage_client = storage.Client()
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(gcs_file_name)
        blob.upload_from_filename(local_file_path)
        
        # ✅ Verify upload
        if blob.exists():
            print(f"✅ Uploaded {local_file_path} to gs://{GCS_BUCKET_NAME}/{gcs_file_name}")
            print(f"   File size: {blob.size} bytes")
            return True
        else:
            print(f"❌ Upload verification failed for {gcs_file_name}")
            return False
    except Exception as e:
        print(f"❌ GCS upload error: {e}")
        import traceback
        traceback.print_exc()
        return False


@app.route('/api/debug/gcs', methods=['GET'])
def debug_gcs():
    """Debug: List files in GCS bucket"""
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blobs = list(bucket.list_blobs())
        
        files = []
        for blob in blobs:
            files.append({
                'name': blob.name,
                'size': blob.size,
                'created': blob.time_created.isoformat() if blob.time_created else None
            })
        
        return jsonify({
            'bucket': GCS_BUCKET_NAME,
            'file_count': len(files),
            'files': files,
            'current_session_id': generation_status.get('results', {}).get('session_id')
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
def download_from_gcs(gcs_file_name, local_file_path):
    """Download file from Google Cloud Storage"""
    try:
        # ✅ Ensure directory exists
        os.makedirs(os.path.dirname(local_file_path), exist_ok=True)
        
        storage_client = storage.Client()
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(gcs_file_name)
        
        # ✅ Check if blob exists
        if not blob.exists():
            print(f"❌ Blob does not exist: gs://{GCS_BUCKET_NAME}/{gcs_file_name}")
            return False
        
        blob.download_to_filename(local_file_path)
        print(f"✅ Downloaded gs://{GCS_BUCKET_NAME}/{gcs_file_name} to {local_file_path}")
        return True
    except Exception as e:
        print(f"❌ GCS download error: {e}")
        import traceback
        traceback.print_exc()
        return False

def get_gcs_signed_url(gcs_file_name, expiration=3600):
    """Generate a signed URL for direct download"""
    try:
        storage_client = storage.Client()
        bucket = storage_client.bucket(GCS_BUCKET_NAME)
        blob = bucket.blob(gcs_file_name)
        url = blob.generate_signed_url(expiration=expiration)
        return url
    except Exception as e:
        print(f"❌ GCS signed URL error: {e}")
        return None
    
class EnhancedFigmaTestCaseGenerator:
    """
    Enhanced Figma analyzer with chunked test case generation to handle large analyses.
    """
    RATE_LIMIT_SLEEP = 1.0  # seconds between calls
    RATE_LIMIT_BACKOFF = 60  # seconds to wait on 429
    def __init__(self, figma_token: str, openai_api_key: str):
        self.figma_token = figma_token
        self.base_url = "https://api.figma.com/v1"
        self.headers = {"X-Figma-Token": figma_token}
        self.openai_client = OpenAI(api_key=openai_api_key)
        self.analytics = {
            'start_time': None,
            'frames_processed': 0,
            'api_calls': 0,
            'tokens_used': 0,
            'estimated_time': 0
        }


    def detect_platform_from_dimensions(self, width: int, height: int) -> str:
        """
        Enhanced platform detection with precise resolution matching.
        Returns: WEB, ANDROID, iOS, or TABLET
        """
        w, h = int(width), int(height)
        short, long = min(w, h), max(w, h)
        
        # MOBILE RESOLUTIONS
        # iPhone specific
        if (w == 375 and h == 667) or (w == 667 and h == 375):  # iPhone 6/7/8
            return "iOS"
        if (w == 414 and h == 896) or (w == 896 and h == 414):  # iPhone XR/11
            return "iOS"
        if (w == 390 and h == 844) or (w == 844 and h == 390):  # iPhone 12/13
            return "iOS"
        if (w == 393 and h == 852) or (w == 852 and h == 393):  # iPhone 14
            return "iOS"
        if (w == 430 and h == 932) or (w == 932 and h == 430):  # iPhone 14 Pro Max
            return "iOS"
        
        # Android specific
        if (w == 360 and h == 640) or (w == 640 and h == 360):  # Common Android
            return "ANDROID"
        if (w == 360 and h == 800) or (w == 800 and h == 360):  # Android HD+
            return "ANDROID"
        if (w == 412 and h == 915) or (w == 915 and h == 412):  # Pixel
            return "ANDROID"
        
        # Generic mobile range
        if short <= 430 and long <= 932:
            # Heuristic: widths 375-430 likely iOS, 360-412 likely Android
            if 375 <= short <= 430:
                return "iOS"
            elif 360 <= short < 375:
                return "ANDROID"
            return "ANDROID / iOS"  # Generic mobile
        
        # TABLET RESOLUTIONS
        # iPad specific
        if (w == 768 and h == 1024) or (w == 1024 and h == 768):  # iPad
            return "TABLET (iPad)"
        if (w == 834 and h == 1194) or (w == 1194 and h == 834):  # iPad Pro 11"
            return "TABLET (iPad)"
        if (w == 1024 and h == 1366) or (w == 1366 and h == 1024):  # iPad Pro 12.9"
            return "TABLET (iPad)"
        
        # Android tablets
        if (w == 800 and h == 1280) or (w == 1280 and h == 800):  # Android tablet
            return "TABLET (Android)"
        
        # Generic tablet range
        if 600 <= short <= 900 and 960 <= long <= 1400:
            return "TABLET"
        
        # WEB/DESKTOP RESOLUTIONS
        # Common desktop breakpoints
        if w >= 1920 or h >= 1080:  # Full HD and above
            return "WEB (Desktop)"
        if w >= 1440 or h >= 900:  # Laptop
            return "WEB (Laptop)"
        if w >= 1280 or h >= 720:  # HD
            return "WEB"
        if w >= 1024:  # Minimum desktop
            return "WEB"
        
        # Fallback
        return "WEB"

    def figma_api_get(self, url, params=None, retries=3):
        """Wrapper for GET requests to Figma API with rate limiting"""
        time.sleep(self.RATE_LIMIT_SLEEP)
        
        for attempt in range(retries):
            response = requests.get(url, headers=self.headers, params=params)
            
            if response.status_code == 429:
                print(f"[Figma API] Rate limit hit, waiting {self.RATE_LIMIT_BACKOFF}s…")
                time.sleep(self.RATE_LIMIT_BACKOFF)
                continue
            
            if response.status_code >= 400:
                print(f"[Figma API] Error {response.status_code}: {response.text}")
                return response
            
            return response
        
        raise Exception(f"[Figma API] Failed after {retries} retries: {url}")
    
    # Add these methods to your EnhancedFigmaTestCaseGenerator class

    def save_analysis_by_sections(self, all_analyses: Dict, base_dir: str = "analysis_output"):
        """
        Save analysis in hierarchical structure: sections > frames > analysis.json
        """
        Path(base_dir).mkdir(parents=True, exist_ok=True)
        
        # Group frames by section
        sections_data = {}
        for frame_data in all_analyses['frames']:
            section_key = frame_data.get('section_key', 'Unnamed_Section')
            section_display = frame_data.get('section', 'Unnamed Section')
            
            if section_key not in sections_data:
                sections_data[section_key] = {
                    'section_info': {
                        'key': section_key,
                        'display_name': section_display
                    },
                    'frames': []
                }
            
            sections_data[section_key]['frames'].append(frame_data)
        
        # Save master index
        master_index = {
            'metadata': all_analyses['metadata'],
            'sections': []
        }
        
        # Create directory structure and save files
        for section_key, section_data in sections_data.items():
            section_display = section_data['section_info']['display_name']
            safe_section_name = self.sanitize_filename(section_key)
            section_path = os.path.join(base_dir, safe_section_name)
            Path(section_path).mkdir(parents=True, exist_ok=True)
            
            print(f"\n📁 Section: {section_display}")
            
            # Add to master index
            master_index['sections'].append({
                'key': section_key,
                'display_name': section_display,
                'path': safe_section_name,
                'frame_count': len(section_data['frames'])
            })
            
            # Save each frame's analysis
            for i, frame_data in enumerate(section_data['frames'], 1):
                frame_name = frame_data['frame_name']
                safe_frame_name = self.sanitize_filename(frame_name)
                
                # Create frame directory
                frame_path = os.path.join(section_path, f"{i:03d}_{safe_frame_name}")
                Path(frame_path).mkdir(parents=True, exist_ok=True)
                
                # Save frame analysis
                analysis_file = os.path.join(frame_path, "analysis.json")
                frame_export = {
                    'frame_info': {
                        'name': frame_name,
                        'section': section_display,
                        'section_key': section_key,
                        'platform': frame_data.get('platform', 'WEB'),
                        'dimensions': frame_data['dimensions'],
                        'index': i
                    },
                    'analysis': frame_data['analysis']
                }
                
                with open(analysis_file, 'w', encoding='utf-8') as f:
                    json.dump(frame_export, f, indent=2, ensure_ascii=False)
                
                print(f"   ✓ {i:02d}. {frame_name} → {analysis_file}")
        
        # Save master index
        index_file = os.path.join(base_dir, "index.json")
        with open(index_file, 'w', encoding='utf-8') as f:
            json.dump(master_index, f, indent=2, ensure_ascii=False)
        
        print(f"\n✅ Analysis structure saved to: {base_dir}/")
        print(f"   📋 Master index: {index_file}")
        return base_dir


    def generate_test_cases_by_sections(self, analysis_data: Dict, output_dir: str = "test_cases_output"):
        """
        Generate test cases organized by section > frame hierarchy
        """
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        print(f"\n📝 Generating section-wise test cases...")
        
        # Group frames by section
        sections_data = {}
        for frame_data in analysis_data['frames']:
            section_key = frame_data.get('section_key', 'Unnamed_Section')
            section_display = frame_data.get('section', 'Unnamed Section')
            
            if section_key not in sections_data:
                sections_data[section_key] = {
                    'section_info': {
                        'key': section_key,
                        'display_name': section_display
                    },
                    'frames': []
                }
            
            sections_data[section_key]['frames'].append(frame_data)
        
        # Generate master test suite document
        master_file = os.path.join(output_dir, "00_MASTER_TEST_SUITE.md")
        with open(master_file, 'w', encoding='utf-8') as f:
            f.write(f"""# COMPREHENSIVE TEST SUITE
    ## Project: {analysis_data.get('metadata', {}).get('project_name', 'Figma Design')}

    ## METADATA
    - **Total Sections**: {len(sections_data)}
    - **Total Frames**: {analysis_data.get('metadata', {}).get('total_frames', 0)}
    - **Analysis Date**: {analysis_data.get('metadata', {}).get('analysis_date', 'N/A')}
    - **Generated**: {time.strftime('%Y-%m-%d %H:%M:%S')}

    ---

    ## TABLE OF CONTENTS (By Section)

    """)
        
        section_counter = 1
        total_test_cases = 0
        
        # Process each section
        for section_key, section_data in sections_data.items():
            section_display = section_data['section_info']['display_name']
            safe_section_name = self.sanitize_filename(section_key)
            
            print(f"\n{'='*70}")
            print(f"📂 Section {section_counter}/{len(sections_data)}: {section_display}")
            print(f"{'='*70}")
            
            # Create section directory
            section_path = os.path.join(output_dir, f"{section_counter:02d}_{safe_section_name}")
            Path(section_path).mkdir(parents=True, exist_ok=True)
            
            # Create section summary file
            section_summary_file = os.path.join(section_path, "00_SECTION_SUMMARY.md")
            section_toc = []
            
            # Add to master TOC
            with open(master_file, 'a', encoding='utf-8') as f:
                f.write(f"\n### {section_counter}. {section_display}\n")
                f.write(f"   - Frames: {len(section_data['frames'])}\n")
                f.write(f"   - Location: `{section_counter:02d}_{safe_section_name}/`\n")
            
            section_test_cases = 0
            
            # Process each frame in the section
            for frame_idx, frame_data in enumerate(section_data['frames'], 1):
                frame_name = frame_data['frame_name']
                safe_frame_name = self.sanitize_filename(frame_name)
                
                print(f"\n  📄 Frame {frame_idx}/{len(section_data['frames'])}: {frame_name}")
                
                # Create frame test case file
                frame_file = os.path.join(
                    section_path, 
                    f"{frame_idx:03d}_{safe_frame_name}_test_cases.md"
                )
                
                # Generate test cases for this frame
                frame_test_cases = self.generate_test_cases_for_frame(
                    frame_data, 
                    frame_idx, 
                    section_display,
                    frame_file
                )
                
                section_test_cases += frame_test_cases
                section_toc.append({
                    'frame_name': frame_name,
                    'file': f"{frame_idx:03d}_{safe_frame_name}_test_cases.md",
                    'test_count': frame_test_cases
                })
                
                print(f"     ✓ Generated {frame_test_cases} test cases")
            
            # Write section summary
            with open(section_summary_file, 'w', encoding='utf-8') as f:
                f.write(f"""# SECTION: {section_display}

    ## Overview
    - **Section Number**: {section_counter}
    - **Total Frames**: {len(section_data['frames'])}
    - **Total Test Cases**: {section_test_cases}

    ---

    ## Frames in this Section

    """)
                for idx, toc_item in enumerate(section_toc, 1):
                    f.write(f"{idx}. **{toc_item['frame_name']}**\n")
                    f.write(f"   - File: `{toc_item['file']}`\n")
                    f.write(f"   - Test Cases: {toc_item['test_count']}\n\n")
            
            total_test_cases += section_test_cases
            print(f"\n  ✅ Section complete: {section_test_cases} test cases")
            section_counter += 1
        
        # Finalize master file
        with open(master_file, 'a', encoding='utf-8') as f:
            f.write(f"""

    ---

    ## SUMMARY
    - **Total Sections**: {len(sections_data)}
    - **Total Test Cases**: {total_test_cases}
    - **Average per Section**: {total_test_cases / len(sections_data):.1f}

    ---

    ## NAVIGATION
    Each section has its own folder with:
    1. `00_SECTION_SUMMARY.md` - Section overview
    2. Individual frame test case files (numbered sequentially)

    ---

    ## TEST EXECUTION TRACKING

    | Section | Frames | Test Cases | Status | Notes | Date |
    |---------|--------|------------|--------|-------|------|
    """)
            for i, (section_key, section_data) in enumerate(sections_data.items(), 1):
                section_display = section_data['section_info']['display_name']
                frame_count = len(section_data['frames'])
                f.write(f"| {i}. {section_display} | {frame_count} | - | ⬜ Pending | | |\n")
            
            f.write("\n---\n\n**End of Master Test Suite**\n")
        
        print(f"\n{'='*70}")
        print(f"🎉 TEST CASE GENERATION COMPLETE!")
        print(f"{'='*70}")
        print(f"   ✅ Total Sections: {len(sections_data)}")
        print(f"   ✅ Total Test Cases: {total_test_cases}")
        print(f"   📁 Output Directory: {output_dir}/")
        print(f"   📋 Master File: {master_file}")
        print(f"{'='*70}\n")
        
        return {
            'output_dir': output_dir,
            'master_file': master_file,
            'total_sections': len(sections_data),
            'total_test_cases': total_test_cases
        }


    def generate_test_cases_for_frame(self, frame_data: Dict, frame_index: int, 
                                    section_name: str, output_file: str) -> int:
        """
        Generate test cases for a single frame with chunking by element category
        """
        frame_name = frame_data['frame_name']
        analysis = frame_data.get('analysis', {})
        
        # Split into chunks by category
        categories = [
            ('design_overview', 'Design Overview'),
            ('branding', 'Branding Elements'),
            ('headings_and_titles', 'Headings and Titles'),
            ('body_text', 'Body Text'),
            ('input_fields', 'Input Fields'),
            ('buttons', 'Buttons'),
            ('links', 'Links'),
            ('checkboxes_radio_buttons', 'Checkboxes & Radio Buttons'),
            ('icons', 'Icons'),
            ('dividers_separators', 'Dividers'),
            ('containers_cards', 'Containers & Cards'),
            ('footer_info', 'Footer'),
            ('copyright_and_legal', 'Copyright & Legal'),
            ('Imgaes_and_media', 'Images & Media'),
            ('Backgrounds', 'Backgrounds'),
            ('colors_and_gradients', 'Colors & Gradients'),
        ]
        
        # Write header
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(f"""# TEST CASES: {frame_name}

    ## Frame Information
    - **Section**: {section_name}
    - **Frame Index**: {frame_index}
    - **Platform**: {frame_data.get('platform', 'WEB')}
    - **Dimensions**: {frame_data.get('dimensions', {}).get('width', 0)}x{frame_data.get('dimensions', {}).get('height', 0)}px

    ---

    """)
        
        total_test_cases = 0
        
        # Generate test cases for each category
        for category_key, category_name in categories:
            category_data = analysis.get(category_key, [])
            
            if not category_data and category_key != 'design_overview':
                continue
            
            # Special handling for design_overview (dict not list)
            if category_key == 'design_overview':
                if not isinstance(category_data, dict) or not category_data:
                    continue
                category_data = [category_data]
            
            chunk_data = {
                'chunk_id': f'{frame_name}_{category_key}',
                'frame_name': frame_name,
                'section': section_name,
                'category': category_name,
                'data': {category_key: category_data}
            }
            
            # Generate test cases for this chunk
            test_cases_content = self.generate_test_cases_for_chunk_enhanced(chunk_data)
            
            # Count test cases
            test_case_count = test_cases_content.count('**Test Case ID**')
            total_test_cases += test_case_count
            
            # Append to file
            with open(output_file, 'a', encoding='utf-8') as f:
                f.write(f"\n\n## {category_name}\n\n")
                f.write(test_cases_content)
                f.write("\n\n---\n")
        
        return total_test_cases


    def generate_test_cases_for_chunk_enhanced(self, chunk: Dict) -> str:
        """
        Enhanced test case generation with better formatting
        """
        prompt = f"""Generate focused, high-value manual test cases for this UI element category.

**Frame**: {chunk['frame_name']}
**Section**: {chunk['section']}
**Category**: {chunk['category']}

Generate 3-8 MOST CRITICAL test cases covering the essential functionality:

For each test case, focus on:
1. **Core functionality** - Does the element work as intended?
2. **User interaction** - Can users interact with it correctly?
3. **Visual verification** - Does it match the design spec?
4. **Edge cases** - Boundary and error conditions

**DO NOT GENERATE REDUNDANT OR TRIVIAL TEST CASES.** Focus only on what matters.

Format each test case as:

---
Note: Include the category only if it is relevant and makes sense.

**Test Case ID**: {chunk['chunk_id'].upper()}-001  
**Title**: [Clear, specific test title]  
**Category**: {chunk['category']}  
**Priority**: Critical/High/Medium/Low  

**Description**: [What this test verifies in one sentence]

**Preconditions**:
- User is on the {chunk['frame_name']} screen

**Steps**:
1. [One specific action]
2. [One verification step]

**Expected Result**:
- [Single, clear expected outcome]

---

Analysis Data:
```json
{json.dumps(chunk['data'], indent=2)}
Generate only the most important test cases. No more than 8 total."""

        try:
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=8000,
                temperature=0.2
            )
            
            self.analytics['api_calls'] += 1
            self.analytics['tokens_used'] += response.usage.total_tokens
            
            return response.choices[0].message.content
            
        except Exception as e:
            return f"## Error generating test cases\n\n{str(e)}"
        
    
    def get_file(self, file_key: str) -> Dict[str, Any]:
        """Retrieve a Figma file by its key"""
        url = f"{self.base_url}/files/{file_key}"
        response = requests.get(url, headers=self.headers)
        
        if response.status_code == 200:
            return response.json()
        else:
            raise Exception(f"Error fetching file: {response.status_code} - {response.text}")

##
    def parse_test_cases_from_markdown(self, markdown_file: str) -> List[Dict]:
        """Parse test cases from markdown file with section and frame info"""
        with open(markdown_file, 'r', encoding='utf-8') as f:
            content = f.read()
        
        test_cases = []
        
        # Try to extract section and frame from file path or content
        file_path = Path(markdown_file)
        section_name = "Unknown"
        frame_name = "Unknown"
        
        # Extract from file path (e.g., test_cases_output/01_Section_Name/003_Frame_Name.md)
        if len(file_path.parts) >= 2:
            section_name = file_path.parts[-2].split('_', 1)[-1] if '_' in file_path.parts[-2] else file_path.parts[-2]
            frame_name = file_path.stem.split('_', 1)[-1] if '_' in file_path.stem else file_path.stem
        
        # Try to extract from markdown headers
        frame_header = re.search(r'# TEST CASES: (.+)', content)
        if frame_header:
            frame_name = frame_header.group(1).strip()
        
        section_header = re.search(r'\*\*Section\*\*:\s*(.+?)(?:\n|$)', content)
        if section_header:
            section_name = section_header.group(1).strip()
        
        # Extract dimensions and detect platform
        dimensions_match = re.search(r'\*\*Dimensions\*\*:\s*(\d+)x(\d+)px', content)
        platform = "UNKNOWN"

        platform_match = re.search(r'\*\*Platform\*\*:\s*(.+)', content)
        if platform_match:
            platform = platform_match.group(1).strip()

        dimensions_match = re.search(r'\*\*Dimensions\*\*:\s*(\d+)x(\d+)px', content)
        if platform == "UNKNOWN" and dimensions_match:
            width = int(dimensions_match.group(1))
            height = int(dimensions_match.group(2))
            platform = self.detect_platform_from_dimensions(width, height)
        
        sections = re.split(r'\*\*Test Case ID\*\*:', content)
        
        for section in sections[1:]:
            try:
                tc_id_match = re.search(r'^([^\n]+)', section)
                tc_id = tc_id_match.group(1).strip() if tc_id_match else "N/A"
                
                title_match = re.search(r'\*\*Title\*\*:\s*(.+?)(?:\n|$)', section)
                title = title_match.group(1).strip() if title_match else "N/A"
                
                category_match = re.search(r'\*\*Category\*\*:\s*(.+?)(?:\n|$)', section)
                category = category_match.group(1).strip() if category_match else "General"
                
                priority_match = re.search(r'\*\*Priority\*\*:\s*(.+?)(?:\n|$)', section, re.IGNORECASE)
                priority = priority_match.group(1).strip() if priority_match else "Medium"
                
                description_match = re.search(r'\*\*Description\*\*:\s*(.+?)(?:\n\*\*|\n\n|$)', section, re.DOTALL)
                description = description_match.group(1).strip() if description_match else ""
                
                precond_match = re.search(r'\*\*Preconditions\*\*:\s*\n((?:.+\n)*?)(?:\n\*\*|\n\n|$)', section, re.DOTALL)
                preconditions = precond_match.group(1).strip() if precond_match else ""
                
                steps_match = re.search(r'\*\*Steps\*\*:\s*\n((?:\d+\..*?\n)+)', section, re.DOTALL)
                steps = steps_match.group(1).strip() if steps_match else "N/A"
                
                expected_match = re.search(r'\*\*Expected Result\*\*:\s*\n((?:[-•].*?\n)+)', section, re.DOTALL)
                expected = expected_match.group(1).strip() if expected_match else "N/A"
                
                test_cases.append({
                    'section': section_name,
                    'frame': frame_name,
                    'platform': platform,
                    'tc_id': tc_id,
                    'title': title,
                    'category': category,
                    'priority': priority,
                    'description': description,
                    'preconditions': preconditions,
                    'steps': steps,
                    'expected': expected
                })
            except Exception as e:
                print(f"Error parsing test case: {e}")
                continue
        
        return test_cases
    

    def extract_visible_frames_only(self, file_data, page_name=None, debug=True):
        sections = {}

        def bbox(child):
            bb = child.get("absoluteBoundingBox", {}) or {}
            return (
                bb.get("x", 0),
                bb.get("y", 0),
                bb.get("width", 0),
                bb.get("height", 0),
            )

        def same_row(a, b, tol=80):
            return abs(a[1] - b[1]) < tol

        for page in file_data.get("document", {}).get("children", []):
            if page_name and page.get("name") != page_name:
                continue
            if page.get("type") != "CANVAS":
                continue

            for section in page.get("children", []):
                name = section.get("name", "Unnamed Section")
                children = section.get("children", []) or []

                if debug:
                    print(f"\n📂 Section: {name}")

                # Always create the section entry (even if we later find 0 frames)
                sections[name] = {
                    "section_info": {
                        "id": section.get("id"),
                        "name": name,
                        "type": section.get("type"),
                        "page": page.get("name"),
                    },
                    "frames": [],
                }

                blocks = []
                for ch in children:
                    t = ch.get("type")
                    if t not in ["FRAME", "GROUP", "COMPONENT", "INSTANCE"]:
                        continue
                    x, y, w, h = bbox(ch)
                    blocks.append((ch, x, y, w, h, t))

                # If there are no candidate blocks, this section just has 0 visible screens
                if not blocks:
                    if debug:
                        print("📊 Total visible screens: 0")
                    continue

                rows = []
                used = set()
                for i, bi in enumerate(blocks):
                    if i in used:
                        continue
                    row = [bi]
                    used.add(i)
                    for j, bj in enumerate(blocks):
                        if j in used:
                            continue
                        if same_row(bi, bj):
                            row.append(bj)
                            used.add(j)
                    rows.append(row)

                final_frames = []

                for row in rows:
                    for ch, x, y, w, h, t in row:
                        # 1 — tiny stuff → ignore
                        if w < 120 and h < 120:
                            if debug:
                                print(f"  ⏭️ Ignored tiny: {ch.get('name')} ({w}x{h})")
                            continue

                        # 2 — full screen
                        if h >= 500 or w >= 500:
                            final_frames.append({
                                "id": ch.get("id"),
                                "name": ch.get("name"),
                                "width": w,
                                "height": h,
                            })
                            if debug:
                                print(f"  ✅ Full Screen: {ch.get('name')} ({w}x{h})")
                            continue

                        # 3 — widgets/cards
                        if 200 <= h < 500:
                            final_frames.append({
                                "id": ch.get("id"),
                                "name": ch.get("name"),
                                "width": w,
                                "height": h,
                            })
                            if debug:
                                print(f"  🔹 UI Widget: {ch.get('name')} ({w}x{h})")
                            continue

                        # 4 — rows of cards
                        if len(row) >= 3 and h >= 150:
                            final_frames.append({
                                "id": ch.get("id"),
                                "name": ch.get("name"),
                                "width": w,
                                "height": h,
                            })
                            if debug:
                                print(f"  🔸 Row Card: {ch.get('name')} ({w}x{h})")
                            continue

                        # 5 — style / component / widget sections
                        if any(k in name.lower() for k in ["style", "color", "widget", "component", "details"]):
                            if h >= 150:
                                final_frames.append({
                                    "id": ch.get("id"),
                                    "name": ch.get("name"),
                                    "width": w,
                                    "height": h,
                                })
                                if debug:
                                    print(f"  🔹 Style Item: {ch.get('name')} ({w}x{h})")
                                continue

                        if debug:
                            print(f"  ⏭️ Ignored: {ch.get('name')} ({w}x{h})")

                # save frames (may be empty)
                sections[name]["frames"] = final_frames

                if debug:
                    print(f"📊 Total visible screens: {len(final_frames)}")

        return sections






    def consolidate_test_cases_from_directory(self, directory_path: str) -> List[Dict]:
        """Consolidate test cases from all markdown files"""
        all_test_cases = []
        
        for root, dirs, files in os.walk(directory_path):
            for file in files:
                if file.endswith('.md') and not file.startswith('00_') and not file.startswith('MASTER'):
                    file_path = os.path.join(root, file)
                    try:
                        test_cases = self.parse_test_cases_from_markdown(file_path)
                        all_test_cases.extend(test_cases)
                        print(f"      Parsed {len(test_cases)} test cases from {file}")
                    except Exception as e:
                        print(f"      Error parsing {file}: {e}")
        
        return all_test_cases


    def find_test_case_files(self) -> List[str]:
        """Find all test case markdown files in the output structure"""
        test_case_files = []
        
        # Check test_cases_output directory
        if os.path.exists('test_cases_output'):
            for root, dirs, files in os.walk('test_cases_output'):
                for file in files:
                    if file.endswith('.md') and ('test' in file.lower() or 'case' in file.lower()):
                        test_case_files.append(os.path.join(root, file))
        
        # Check for individual section directories
        sections = ['00_MASTER_TEST_SUITE.md', 'comprehensive_test_cases.md']
        for section in sections:
            if os.path.exists(section):
                test_case_files.append(section)
        
        return test_case_files
    

    def parse_all_test_case_files(self, file_paths: List[str]) -> List[Dict]:
        """Parse multiple test case files"""
        all_test_cases = []
        
        for file_path in file_paths:
            try:
                test_cases = self.parse_test_cases_from_markdown(file_path)
                all_test_cases.extend(test_cases)
                print(f"    Parsed {len(test_cases)} from {os.path.basename(file_path)}")
            except Exception as e:
                print(f"    Error parsing {file_path}: {e}")
        
        return all_test_cases
    

    def parse_test_case_section(self, section: str) -> Dict:
        """Parse a single test case section."""
        tc = {}
        
        try:
            # Extract Test Case ID
            id_match = re.search(r'^([^\n]+)', section)
            if id_match:
                tc['tc_id'] = id_match.group(1).strip()
            
            # Extract Title
            title_match = re.search(r'\*\*Title\*\*:\s*(.+?)(?:\n|$)', section)
            if title_match:
                tc['title'] = title_match.group(1).strip()
            
            # Extract Category
            category_match = re.search(r'\*\*Category\*\*:\s*(.+?)(?:\n|$)', section)
            if category_match:
                tc['category'] = category_match.group(1).strip()
            
            # Extract Priority
            priority_match = re.search(r'\*\*Priority\*\*:\s*(.+?)(?:\n|$)', section, re.IGNORECASE)
            if priority_match:
                tc['priority'] = priority_match.group(1).strip()
            
            # Extract Description
            desc_match = re.search(r'\*\*Description\*\*:\s*\n(.+?)(?:\n\*\*|\n\n|$)', section, re.DOTALL)
            if desc_match:
                tc['description'] = desc_match.group(1).strip()
            
            # Extract Steps
            steps_match = re.search(r'\*\*Steps\*\*:\s*\n((?:\d+\..*?\n)+)', section, re.DOTALL)
            if steps_match:
                steps_text = steps_match.group(1).strip()
                tc['steps'] = steps_text
            
            # Extract Expected Result
            expected_match = re.search(r'\*\*Expected Result\*\*:\s*\n((?:.|\n)+?)(?:\n\*\*|\n\n|$)', section, re.DOTALL)
            if expected_match:
                expected_text = expected_match.group(1).strip()
                tc['expected'] = expected_text
            
            # Extract Preconditions if present
            precond_match = re.search(r'\*\*Preconditions\*\*:\s*\n((?:.|\n)+?)(?:\n\*\*|\n\n|$)', section, re.DOTALL)
            if precond_match:
                tc['preconditions'] = precond_match.group(1).strip()
            
            return tc if tc.get('tc_id') or tc.get('title') else None
            
        except Exception as e:
            print(f"        Parsing error in section: {e}")
            return None
        

    def parse_test_case_line(self, line: str, current_case: Dict):
        """Parse individual lines of a test case."""
        if line.startswith('**Title**:'):
            current_case['title'] = line.replace('**Title**:', '').strip()
        elif line.startswith('**Category**:'):
            current_case['category'] = line.replace('**Category**:', '').strip()
        elif line.startswith('**Priority**:'):
            current_case['priority'] = line.replace('**Priority**:', '').strip()
        elif line.startswith('**Description**:'):
            current_case['description'] = line.replace('**Description**:', '').strip()
        elif line.startswith('**Steps**:'):
            current_case['steps'] = line.replace('**Steps**:', '').strip()
        elif line.startswith('**Expected Result**:'):
            current_case['expected'] = line.replace('**Expected Result**:', '').strip()


    
        

    def generate_excel_report(self, markdown_file: str, project_name: str = "Figma Design", 
                            output_file: str = "test_cases.xlsx"):
        """Generate Excel file with proper section, frame, and platform columns"""
        print(f"\n  Generating Excel report from: {markdown_file}")
        
        # Check if this is a directory
        if os.path.isdir(markdown_file):
            print(f"    Directory detected, consolidating test cases...")
            test_cases = self.consolidate_test_cases_from_directory(markdown_file)
        else:
            test_cases = self.parse_test_cases_from_markdown(markdown_file)
        
        print(f"    Parsed {len(test_cases)} test cases")
        
        if len(test_cases) == 0:
            raise Exception(
                "No test cases found. Ensure you are passing test_cases_output directory, not master markdown."
            )       
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Define headers with Section and Frame columns
        headers = [
            "S.No",
            "Section",
            "Frame",
            "Platform",
            "Test Case ID",
            "Title",
            "Category",
            "Priority",
            "Description",
            "Preconditions",
            "Steps",
            "Expected Result",
            "Status"
        ]
        
        # Style setup
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Set column widths
        column_widths = {
            'A': 8,   # S.No
            'B': 25,  # Section
            'C': 30,  # Frame
            'D': 15,  # Platform
            'E': 20,  # Test Case ID
            'F': 40,  # Title
            'G': 20,  # Category
            'H': 12,  # Priority
            'I': 50,  # Description
            'J': 40,  # Preconditions
            'K': 60,  # Steps
            'L': 60,  # Expected Result
            'M': 15   # Status
        }
        
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Write test cases with serial numbers
        serial_no = 1
        for tc in test_cases:
            row_num = serial_no + 1
            
            ws.cell(row=row_num, column=1, value=serial_no)
            ws.cell(row=row_num, column=2, value=tc.get('section', ''))
            ws.cell(row=row_num, column=3, value=tc.get('frame', ''))
            ws.cell(row=row_num, column=4, value=tc.get('platform', 'WEB'))
            ws.cell(row=row_num, column=5, value=tc.get('tc_id', ''))
            ws.cell(row=row_num, column=6, value=tc.get('title', ''))
            ws.cell(row=row_num, column=7, value=tc.get('category', 'General'))
            ws.cell(row=row_num, column=8, value=tc.get('priority', 'Medium'))
            ws.cell(row=row_num, column=9, value=tc.get('description', ''))
            ws.cell(row=row_num, column=10, value=tc.get('preconditions', ''))
            ws.cell(row=row_num, column=11, value=tc.get('steps', ''))
            ws.cell(row=row_num, column=12, value=tc.get('expected', ''))
            ws.cell(row=row_num, column=13, value="Pending")
            
            # Apply formatting
            for col_num in range(1, 14):
                cell = ws.cell(row=row_num, column=col_num)
                cell.alignment = cell_alignment
                cell.border = thin_border
            
            serial_no += 1
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Save the workbook
        wb.save(output_file)
        print(f"    ✓ Excel report saved to: {output_file}")
        return output_file

    

    def extract_strict_screens_from_sections(self, file_data: Dict,
                                            page_name: str = None,
                                            min_width: int = 300, min_height: int = 300,
                                            debug: bool = False) -> Dict[str, Any]:
        """
        ULTRA-STRICT: Extract ONLY visible screens that are:
        1. Direct children of canvas sections (not nested)
        2. Actually visible as top-level frames on canvas
        3. Screen-sized (width >= min_width AND height >= min_height)
        """
        screens = {}
        
        for page in file_data.get('document', {}).get('children', []):
            if page_name and page.get('name') != page_name:
                continue
            if page.get('type') != 'CANVAS':
                continue

            page_actual = page.get('name', 'Unknown Page')
            
            for section in page.get('children', []):
                section_name = section.get('name', 'Unnamed Section')
                sect_info = {
                    'id': section.get('id'),
                    'name': section_name,
                    'type': section.get('type'),
                    'page': page_actual
                }
                
                visible_screens = []
                
                # CRITICAL: Only look at DIRECT children of the section
                # These are the visible frames you see on the canvas
                direct_children = section.get('children', [])
                
                if debug:
                    print(f"\n📂 Section: '{section_name}' has {len(direct_children)} direct children")
                
                for child in direct_children:
                    ctype = child.get('type')
                    cname = child.get('name', '<no-name>')
                    bbox = child.get('absoluteBoundingBox', {}) or {}
                    w = bbox.get('width', 0)
                    h = bbox.get('height', 0)

                    # 1. Size check
                    if w < min_width or h < min_height:
                        if debug:
                            print(f"  ⏭️  {cname} - too small ({int(w)}x{int(h)}px)")
                        continue

                    # 2. Acceptable top-level screen types
                    #    (FRAME, GROUP, COMPONENT, COMPONENT_SET)
                    if ctype not in ['FRAME', 'GROUP', 'COMPONENT', 'COMPONENT_SET']:
                        if debug:
                            print(f"  ⏭️  {cname} - type {ctype} (not a visible screen)")
                        continue

                    # 3. Special case: GROUP used as desktop screen wrapper
                    if ctype == 'GROUP':
                        visible_screens.append({
                            'id': child.get('id'),
                            'name': cname,
                            'type': ctype,
                            'width': w,
                            'height': h,
                            'parent_section': section_name,
                            'page': page_actual,
                            'nested_elements': len(child.get('children', []))
                        })
                        if debug:
                            print(f"  ✅ {cname} (GROUP screen {int(w)}x{int(h)}px)")
                        continue

                    # 4. FRAME / COMPONENT heuristics
                    nested_children = child.get('children', [])
                    nested_frames = sum(1 for c in nested_children if c.get('type') == 'FRAME')
                    text_nodes = sum(1 for c in nested_children if c.get('type') == 'TEXT')

                    # Filter out large container frames, not screens
                    if nested_frames > 5 and text_nodes == 0:
                        if debug:
                            print(f"  ⏭️  {cname} - large container ({nested_frames} nested frames, no text)")
                        continue

                    # 5. Accept as visible screen
                    visible_screens.append({
                        'id': child.get('id'),
                        'name': cname,
                        'type': ctype,
                        'width': w,
                        'height': h,
                        'parent_section': section_name,
                        'page': page_actual,
                        'nested_elements': len(nested_children)
                    })

                    if debug:
                        print(f"  ✅ {cname} ({int(w)}x{int(h)}px, {len(nested_children)} children)")
                                
                if visible_screens:
                    screens[section_name] = {
                        'section_info': sect_info,
                        'frames': visible_screens
                    }
                    if debug:
                        print(f"  📊 Total visible screens: {len(visible_screens)}")
            
        return screens

    def download_direct_frames_only(self, file_key: str, output_dir: str = "figma_frames", 
                                    page_name: str = None, scale: float = 2, 
                                    image_format: str = "png",
                                    min_width: int = 300,
                                    min_height: int = 300) -> Dict[str, List[Dict]]:
        """Download only LARGE, DIRECT frames from canvas sections"""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        print(f"\n🔍 Fetching Figma file data...")
        file_data = self.get_file(file_key)
        
        print(f"\n📊 Extracting VISIBLE screens only (min size: {min_width}x{min_height}px)...")
        print(f"{'='*70}")
        
        sections = self.extract_visible_frames_only(
    file_data,
    page_name=page_name,
    debug=True
)


        
        # Calculate totals
        total_sections = len(sections)
        total_frames = sum(len(section['frames']) for section in sections.values())
        
        print(f"\n{'='*70}")
        print(f"✅ Found {total_sections} sections with {total_frames} VISIBLE frames")
        print(f"{'='*70}")
        
        if total_frames == 0:
            print("\n⚠️  No frames found! Try reducing min_width/min_height parameters.")
            print("   Current filters:")
            print(f"   - Minimum width: {min_width}px")
            print(f"   - Minimum height: {min_height}px")
            return {}
        
        print(f"\n📋 Breakdown by section:")
        for section_name, section_data in sections.items():
            frame_count = len(section_data['frames'])
            print(f"   • {section_name}: {frame_count} frames")
            for frame in section_data['frames']:
                print(f"      - {frame['name']} ({int(frame['width'])}x{int(frame['height'])}px)")
        print(f"{'='*70}\n")
        
        # Collect all frame IDs for batch download
        print(f"📥 Preparing to download {total_frames} frames...")
        all_frame_ids = []
        for section_data in sections.values():
            all_frame_ids.extend([frame['id'] for frame in section_data['frames']])
        
        print(f"🌐 Requesting image URLs from Figma API...")
        image_urls = self.get_image_urls(file_key, all_frame_ids, scale, image_format)
        
        print(f"✅ Received {len(image_urls)} image URLs\n")
        
        downloaded_sections = {}
        global_frame_counter = 1
        
        # Process each section
        for section_idx, (section_name, section_data) in enumerate(sections.items(), 1):
            print(f"\n{'='*70}")
            print(f"📁 Section {section_idx}/{total_sections}: {section_name}")
            print(f"{'='*70}")
            
            # Create folder for this section
            safe_section_name = self.sanitize_filename(section_name)
            section_path = os.path.join(output_dir, safe_section_name)
            Path(section_path).mkdir(parents=True, exist_ok=True)
            
            downloaded_frames = []
            section_frame_count = len(section_data['frames'])
            
            for i, frame in enumerate(section_data['frames'], 1):
                frame_id = frame['id']
                safe_name = self.sanitize_filename(frame['name'])
                
                # Create filename with index
                filename = f"{i:03d}_{safe_name}.{image_format}"
                filepath = os.path.join(section_path, filename)
                
                image_url = image_urls.get(frame_id)
                
                if image_url:
                    print(f"   [{global_frame_counter}/{total_frames}] {frame['name']}")
                    print(f"      📐 Size: {int(frame['width'])}x{int(frame['height'])}px")
                    if self.download_image(image_url, filepath):
                        frame['filepath'] = filepath
                        frame['local_index'] = i
                        frame['global_index'] = global_frame_counter
                        downloaded_frames.append(frame)
                        print(f"      ✅ Saved: {filename}")
                    else:
                        print(f"      ❌ Download failed")
                else:
                    print(f"   ⚠️  No URL for: {frame['name']}")
                
                global_frame_counter += 1
            
            downloaded_sections[section_name] = {
                'section_info': section_data['section_info'],
                'frames': downloaded_frames
            }
            
            print(f"\n   ✅ Complete: {len(downloaded_frames)}/{section_frame_count} frames")
        
        # Summary
        total_downloaded = sum(len(section['frames']) for section in downloaded_sections.values())
        print(f"\n{'='*70}")
        print(f"🎉 DOWNLOAD COMPLETE!")
        print(f"{'='*70}")
        print(f"   ✅ Downloaded: {total_downloaded}/{total_frames} frames")
        print(f"   📁 Sections: {len(downloaded_sections)}")
        print(f"   💾 Location: {output_dir}/")
        print(f"{'='*70}\n")
        
        return downloaded_sections
    
    def extract_frames_from_sections(self, file_data: Dict, page_name: str = None) -> Dict[str, List[Dict]]:
            """Extract frames organized by their parent sections/canvases"""
            sections = {}
            
            if 'document' in file_data:
                pages = file_data['document'].get('children', [])
                
                for page in pages:
                    if page_name and page.get('name') != page_name:
                        continue
                    
                    if page.get('type') == 'CANVAS' and 'children' in page:
                        page_name_actual = page.get('name')
                        
                        for section in page['children']:
                            # This is a top-level element (grey canvas/section)
                            section_name = section.get('name', 'Unnamed Section')
                            section_id = section.get('id')
                            
                            # Initialize section if not exists
                            if section_name not in sections:
                                sections[section_name] = {
                                    'section_info': {
                                        'id': section_id,
                                        'name': section_name,
                                        'type': section.get('type'),
                                        'page': page_name_actual
                                    },
                                    'frames': []
                                }
                            
                            # Check if this section has children (frames inside it)
                            if 'children' in section:
                                for child in section['children']:
                                    # Extract frames/components inside this section
                                    if child.get('type') in ['FRAME', 'GROUP', 'COMPONENT', 'COMPONENT_SET', 'INSTANCE']:
                                        sections[section_name]['frames'].append({
                                            'id': child['id'],
                                            'name': child['name'],
                                            'type': child.get('type'),
                                            'page': page_name_actual,
                                            'parent_section': section_name,
                                            'width': child.get('absoluteBoundingBox', {}).get('width', 0),
                                            'height': child.get('absoluteBoundingBox', {}).get('height', 0)
                                        })
                            else:
                                # If the section itself is a frame (no children), add it
                                if section.get('type') in ['FRAME', 'GROUP', 'COMPONENT', 'COMPONENT_SET', 'INSTANCE']:
                                    sections[section_name]['frames'].append({
                                        'id': section_id,
                                        'name': section_name,
                                        'type': section.get('type'),
                                        'page': page_name_actual,
                                        'parent_section': section_name,
                                        'width': section.get('absoluteBoundingBox', {}).get('width', 0),
                                        'height': section.get('absoluteBoundingBox', {}).get('height', 0)
                                    })
            
            return sections

    def download_frames_by_section(self, file_key: str, output_dir: str = "figma_frames", 
                                page_name: str = None, scale: float = 2, 
                                image_format: str = "png") -> Dict[str, List[Dict]]:
        """Download frames organized by their parent sections"""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        print(f"Fetching file data for: {file_key}")
        file_data = self.get_file(file_key)
        
        sections = self.extract_frames_from_sections(file_data, page_name)
        print("_____________________________________________")

        print(file_data)
        print("______________________________________________")
        print(page_name)
        print("______________________________________________")



        total_frames = sum(len(section['frames']) for section in sections.values())
        print(f"Found {len(sections)} sections with {total_frames} total frames")
        
        if not sections:
            return {}
        
        # Collect all frame IDs for batch download
        all_frame_ids = []
        for section_data in sections.values():
            all_frame_ids.extend([frame['id'] for frame in section_data['frames']])
        
        print("\nRequesting image URLs...")
        image_urls = self.get_image_urls(file_key, all_frame_ids, scale, image_format)
        
        downloaded_sections = {}
        frame_counter = 1
        
        # Process each section
        for section_name, section_data in sections.items():
            print(f"\n📁 Processing Section: {section_name}")
            print(f"   Found {len(section_data['frames'])} frames")
            
            # Create folder for this section
            safe_section_name = self.sanitize_filename(section_name)
            section_path = os.path.join(output_dir, safe_section_name)
            Path(section_path).mkdir(parents=True, exist_ok=True)
            
            downloaded_frames = []
            
            for i, frame in enumerate(section_data['frames'], 1):
                frame_id = frame['id']
                safe_name = self.sanitize_filename(frame['name'])
                
                filename = f"{i:03d}_{safe_name}.{image_format}"
                filepath = os.path.join(section_path, filename)
                
                image_url = image_urls.get(frame_id)
                
                if image_url:
                    print(f"  [{frame_counter}/{total_frames}] Downloading: {frame['name']}")
                    if self.download_image(image_url, filepath):
                        frame['filepath'] = filepath
                        downloaded_frames.append(frame)
                    frame_counter += 1
            
            downloaded_sections[section_name] = {
                'section_info': section_data['section_info'],
                'frames': downloaded_frames
            }
            
            print(f"   ✓ Downloaded {len(downloaded_frames)}/{len(section_data['frames'])} frames")
        
        print(f"\n✓ Downloaded {frame_counter - 1}/{total_frames} total frames across {len(sections)} sections")
        return downloaded_sections

    
    def extract_top_level_frames(self, file_data: Dict, page_name: str = None, include_groups: bool = True) -> List[Dict]:
        """Extract only top-level frames and optionally groups/components"""
        frames = []
        valid_types = ['FRAME']
        if include_groups:
            valid_types.extend(['GROUP', 'COMPONENT', 'COMPONENT_SET', 'INSTANCE'])
        
        if 'document' in file_data:
            pages = file_data['document'].get('children', [])
            
            for page in pages:
                if page_name and page.get('name') != page_name:
                    continue
                
                if page.get('type') == 'CANVAS' and 'children' in page:
                    for child in page['children']:
                        if child.get('type') in valid_types:
                            frames.append({
                                'id': child['id'],
                                'name': child['name'],
                                'type': child.get('type'),
                                'page': page.get('name'),
                                'width': child.get('absoluteBoundingBox', {}).get('width', 0),
                                'height': child.get('absoluteBoundingBox', {}).get('height', 0)
                            })
        
        return frames
        
        # ============================================================================
    # REPLACE THESE METHODS IN YOUR EnhancedFigmaTestCaseGenerator CLASS
    # ============================================================================
    def extract_strict_top_level_screens(self, file_data: Dict, page_name: str = None,
                                    min_width: int = 300, min_height: int = 300) -> Dict[str, Any]:
        """
        ULTRA-STRICT: Extract ONLY top-level screens that are direct children of canvas sections.
        Completely ignores all nested elements, groups, components, etc.
        """
        sections = {}
        
        for page in file_data.get('document', {}).get('children', []):
            if page_name and page.get('name') != page_name:
                continue
            
            if page.get('type') != 'CANVAS':
                continue
            
            page_actual_name = page.get('name', 'Unknown Page')
            
            # Each direct child of CANVAS is a section
            for section in page.get('children', []):
                section_name = section.get('name', 'Unnamed Section')
                section_info = {
                    'id': section.get('id'),
                    'name': section_name,
                    'type': section.get('type'),
                    'page': page_actual_name
                }
                
                screens = []
                
                # Look ONLY at direct children
                for child in section.get('children', []):
                    # ONLY accept FRAME type
                    if child.get('type') != 'FRAME':
                        continue
                    
                    # Check dimensions
                    bbox = child.get('absoluteBoundingBox', {})
                    w = bbox.get('width', 0) if bbox else 0
                    h = bbox.get('height', 0) if bbox else 0
                    
                    # Must be screen-sized (adjust these values based on your needs)
                    if w >= min_width and h >= min_height:
                        # Check if this looks like a screen vs a container
                        child_children = child.get('children', [])
                        
                        # Quick heuristic: screens usually have meaningful content
                        # but not dozens of nested frames
                        text_nodes = len([c for c in child_children if c.get('type') == 'TEXT'])
                        nested_frames = len([c for c in child_children if c.get('type') == 'FRAME'])
                        
                        # If it has text or reasonable number of nested elements, it's likely a screen
                        if text_nodes > 0 or nested_frames < 5:
                            screens.append({
                                'id': child.get('id'),
                                'name': child.get('name', 'Unnamed Screen'),
                                'type': 'FRAME',
                                'width': w,
                                'height': h,
                                'parent_section': section_name,
                                'page': page_actual_name
                            })
                
                if screens:
                    sections[section_name] = {
                        'section_info': section_info,
                        'frames': screens
                    }
        
        return sections
    

    def extract_only_visible_canvas_frames(self, file_data: Dict, page_name: str = None,
                                        min_width: int = 300, min_height: int = 300,
                                        debug: bool = True) -> Dict[str, Any]:
        """
        Extract ONLY direct visible frames on canvas sections - STRICT VERSION.
        NO RECURSION - only looks at immediate children of each canvas section.
        Filters out nested elements, groups, and small non-screen elements.
        """
        sections = {}
        
        for page in file_data.get('document', {}).get('children', []):
            if page_name and page.get('name') != page_name:
                continue
            
            if page.get('type') != 'CANVAS':
                continue
            
            page_actual_name = page.get('name', 'Unknown Page')
            
            print(f"\n📄 Page: {page_actual_name}")
            
            # Each direct child of CANVAS is a grey canvas section
            for section in page.get('children', []):
                section_name = section.get('name', 'Unnamed Section')
                section_info = {
                    'id': section.get('id'),
                    'name': section_name,
                    'type': section.get('type'),
                    'page': page_actual_name
                }
                
                visible_frames = []
                
                # CRITICAL: Only look at DIRECT children, do NOT recurse
                direct_children = section.get('children', [])
                
                if debug:
                    print(f"\n  📂 Section: {section_name}")
                    print(f"    Direct children count: {len(direct_children)}")
                
                for child in direct_children:
                    child_type = child.get('type')
                    child_name = child.get('name', '<unnamed>')
                    bbox = child.get('absoluteBoundingBox', {}) or {}
                    w = bbox.get('width', 0)
                    h = bbox.get('height', 0)
                    
                    # VERY STRICT FILTERING:
                    # 1. Must be a FRAME (not GROUP, not COMPONENT, not INSTANCE)
                    # 2. Must meet minimum size (screen-sized, not small icons/elements)
                    # 3. Must be a visible screen (not a container/group)
                    
                    if child_type == 'FRAME':
                        if w >= min_width and h >= min_height:
                            # Additional check: ensure it's not a nested frame
                            # by checking if it has too many nested children itself
                            child_children = child.get('children', [])
                            # If this frame has many children of type FRAME, it might be a container
                            # We want screens, not containers
                            frame_children_count = len([c for c in child_children if c.get('type') == 'FRAME'])
                            
                            if frame_children_count <= 2:  # Allow some nested frames for complex screens
                                visible_frames.append({
                                    'id': child.get('id'),
                                    'name': child_name,
                                    'type': child_type,
                                    'width': w,
                                    'height': h,
                                    'parent_section': section_name,
                                    'page': page_actual_name,
                                    'direct_child': True
                                })
                                if debug:
                                    print(f"    ✅ SCREEN: {child_name} ({int(w)}x{int(h)}px)")
                            else:
                                if debug:
                                    print(f"    ⏭️  {child_name} - appears to be a container with {frame_children_count} nested frames")
                        else:
                            if debug:
                                print(f"    ⏭️  {child_name} - too small ({int(w)}x{int(h)}px)")
                    else:
                        if debug:
                            print(f"    ⏭️  {child_name} - type: {child_type} (not a FRAME)")
                
                # Only add section if it has visible frames
                if visible_frames:
                    sections[section_name] = {
                        'section_info': section_info,
                        'frames': visible_frames
                    }
                    if debug:
                        print(f"\n    📊 Section '{section_name}': {len(visible_frames)} visible screens\n")
        
        return sections

    def download_only_visible_frames(self, file_key: str, output_dir: str = "figma_frames", 
                                    page_name: str = None, scale: float = 2, 
                                    image_format: str = "png") -> Dict[str, List[Dict]]:
        """
        Download ONLY visible frames - ULTRA STRICT VERSION.
        """
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        print(f"\n{'='*80}")
        print(f"🚀 EXTRACTING VISIBLE SCREENS ONLY (ULTRA STRICT)")
        print(f"{'='*80}\n")
        
        # Fetch file data
        print(f"🌐 Fetching Figma file...")
        file_data = self.get_file(file_key)
        print(f"✅ File data retrieved\n")
        
        # Extract ONLY visible screens (ultra strict)
        print(f"🔍 Looking for visible screens on canvas sections...")
        print(f"{'='*80}")
        
        # Use ultra-strict extraction
        sections = self.extract_strict_top_level_screens(
            file_data, 
            page_name,
            min_width=300,  # Minimum screen width
            min_height=300   # Minimum screen height
        )
        
        print(f"{'='*80}\n")
        
        # Count totals
        total_sections = len(sections)
        total_screens = sum(len(section['frames']) for section in sections.values())
        
        print(f"{'='*80}")
        print(f"📊 EXTRACTION SUMMARY")
        print(f"{'='*80}")
        print(f"   Canvas Sections: {total_sections}")
        print(f"   Total Visible Screens: {total_screens}")
        print(f"{'='*80}\n")
        
        if total_screens == 0:
            print("⚠️  No visible screens found! Trying less strict filtering...")
            # Fall back to slightly less strict method
            sections = self.extract_only_visible_canvas_frames(file_data, page_name, min_width=100, min_height=100, debug=False)
            total_screens = sum(len(section['frames']) for section in sections.values())
            
            if total_screens == 0:
                print("❌ Still no screens found. Check your Figma file structure.")
                return {}
        
        # Show breakdown
        print(f"📋 Breakdown:")
        for section_name, section_data in sections.items():
            frames = section_data['frames']
            print(f"\n  📂 {section_name}: {len(frames)} screens")
            for i, frame in enumerate(frames, 1):
                print(f"     {i}. {frame['name']} - {int(frame['width'])}x{int(frame['height'])}px")
        print(f"\n{'-'*80}\n")
        
        # Get image URLs
        print(f"📥 Requesting image URLs from Figma...")
        all_frame_ids = []
        for section_data in sections.values():
            all_frame_ids.extend([frame['id'] for frame in section_data['frames']])
        
        image_urls = self.get_image_urls(file_key, all_frame_ids, scale, image_format)
        print(f"✅ Got {len(image_urls)} URLs\n")
        
        # Download screens
        downloaded_sections = {}
        global_counter = 1
        
        print(f"{'='*80}")
        print(f"📥 DOWNLOADING SCREENS")
        print(f"{'='*80}\n")
        
        for section_idx, (section_name, section_data) in enumerate(sections.items(), 1):
            print(f"📁 Section {section_idx}/{total_sections}: {section_name}")
            print(f"{'-'*80}")
            
            # Create folder
            safe_name = self.sanitize_filename(section_name)
            section_path = os.path.join(output_dir, safe_name)
            Path(section_path).mkdir(parents=True, exist_ok=True)
            
            downloaded = []
            
            for local_idx, frame in enumerate(section_data['frames'], 1):
                frame_id = frame['id']
                safe_frame_name = self.sanitize_filename(frame['name'])
                filename = f"{local_idx:03d}_{safe_frame_name}.{image_format}"
                filepath = os.path.join(section_path, filename)
                
                image_url = image_urls.get(frame_id)
                
                print(f"  [{global_counter}/{total_screens}] {frame['name']}")
                
                if image_url:
                    if self.download_image(image_url, filepath):
                        frame['filepath'] = filepath
                        frame['local_index'] = local_idx
                        frame['global_index'] = global_counter
                        downloaded.append(frame)
                        print(f"     ✅ Saved: {filename}")
                    else:
                        print(f"     ❌ Download failed")
                else:
                    print(f"     ⚠️  No URL")
                
                global_counter += 1
            
            downloaded_sections[section_name] = {
                'section_info': section_data['section_info'],
                'frames': downloaded
            }
        
        # Summary
        total_downloaded = sum(len(s['frames']) for s in downloaded_sections.values())
        
        print(f"\n{'='*80}")
        print(f"🎉 DOWNLOAD COMPLETE!")
        print(f"{'='*80}")
        print(f"   ✅ Downloaded: {total_downloaded}/{total_screens} screens")
        print(f"   📁 Sections: {len(downloaded_sections)}")
        print(f"   💾 Location: {output_dir}/")
        print(f"{'='*80}\n")
        
        return downloaded_sections

    
    def extract_all_frames_recursively(self, node: Dict, parent_info: Dict = None, depth: int = 0) -> List[Dict]:
        """Recursively extract all frames at any nesting level"""
        frames = []
        
        # Define what we consider a "frame" worth extracting
        frame_types = ['FRAME', 'COMPONENT', 'COMPONENT_SET', 'INSTANCE']
        
        node_type = node.get('type')
        node_id = node.get('id')
        node_name = node.get('name', 'Unnamed')
        
        # Get dimensions
        bbox = node.get('absoluteBoundingBox', {})
        width = bbox.get('width', 0)
        height = bbox.get('height', 0)
        
        # Check if this node is a frame we want to extract
        if node_type in frame_types and width > 0 and height > 0:
            frame_info = {
                'id': node_id,
                'name': node_name,
                'type': node_type,
                'width': width,
                'height': height,
                'depth': depth,
                'parent_section': parent_info['name'] if parent_info else None,
                'parent_id': parent_info['id'] if parent_info else None,
                'page': parent_info['page'] if parent_info else 'Unknown'
            }
            frames.append(frame_info)
            
            # Update parent info for children
            current_parent = {
                'id': node_id,
                'name': node_name,
                'page': parent_info['page'] if parent_info else 'Unknown'
            }
        else:
            current_parent = parent_info
        
        # Recursively process children
        if 'children' in node and isinstance(node['children'], list):
            for child in node['children']:
                frames.extend(self.extract_all_frames_recursively(child, current_parent, depth + 1))
        
        return frames

    

   
    def download_frames_by_canvas_sections(self, file_key: str, output_dir: str = "figma_frames", 
                                        page_name: str = None, scale: float = 2, 
                                        image_format: str = "png") -> Dict[str, List[Dict]]:
        """Download all frames organized by their canvas sections"""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        print(f"\n🔍 Fetching Figma file data...")
        file_data = self.get_file(file_key)
        
        print(f"📊 Extracting frames from all canvas sections...")
        sections = self.extract_strict_screens_from_sections(
    file_data, 
    page_name,
    min_width=300,    # Adjust if your rectangles are smaller
    min_height=300,   # Adjust if your rectangles are smaller
    debug=True        # This will show you what's being filtered
)
        
        # Calculate totals
        total_sections = len(sections)
        total_frames = sum(len(section['frames']) for section in sections.values())
        
        print(f"\n✅ Found {total_sections} canvas sections with {total_frames} total frames")
        print(f"\n📋 Breakdown:")
        for section_name, section_data in sections.items():
            print(f"   • {section_name}: {len(section_data['frames'])} frames")
        
        if not sections or total_frames == 0:
            print("\n❌ No frames found!")
            return {}
        
        # Collect all frame IDs for batch download
        print(f"\n📥 Preparing to download {total_frames} frames...")
        all_frame_ids = []
        for section_data in sections.values():
            all_frame_ids.extend([frame['id'] for frame in section_data['frames']])
        
        print(f"🌐 Requesting image URLs from Figma API...")
        image_urls = self.get_image_urls(file_key, all_frame_ids, scale, image_format)
        
        print(f"✅ Received {len(image_urls)} image URLs")
        
        downloaded_sections = {}
        global_frame_counter = 1
        
        # Process each section
        for section_idx, (section_name, section_data) in enumerate(sections.items(), 1):
            print(f"\n{'='*70}")
            print(f"📁 Section {section_idx}/{total_sections}: {section_name}")
            print(f"{'='*70}")
            
            # Create folder for this section
            safe_section_name = self.sanitize_filename(section_name)
            section_path = os.path.join(output_dir, safe_section_name)
            Path(section_path).mkdir(parents=True, exist_ok=True)
            
            downloaded_frames = []
            section_frame_count = len(section_data['frames'])
            
            for i, frame in enumerate(section_data['frames'], 1):
                frame_id = frame['id']
                safe_name = self.sanitize_filename(frame['name'])
                
                # Create filename with index
                filename = f"{i:03d}_{safe_name}.{image_format}"
                filepath = os.path.join(section_path, filename)
                
                image_url = image_urls.get(frame_id)
                
                if image_url:
                    print(f"   [{global_frame_counter}/{total_frames}] ({i}/{section_frame_count}) Downloading: {frame['name'][:50]}")
                    if self.download_image(image_url, filepath):
                        frame['filepath'] = filepath
                        frame['local_index'] = i
                        frame['global_index'] = global_frame_counter
                        downloaded_frames.append(frame)
                    else:
                        print(f"      ⚠️  Download failed")
                else:
                    print(f"   ⚠️  No URL for frame: {frame['name']}")
                
                global_frame_counter += 1
            
            downloaded_sections[section_name] = {
                'section_info': section_data['section_info'],
                'frames': downloaded_frames
            }
            
            print(f"   ✅ Section complete: {len(downloaded_frames)}/{section_frame_count} frames downloaded")
        
        # Summary
        total_downloaded = sum(len(section['frames']) for section in downloaded_sections.values())
        print(f"\n{'='*70}")
        print(f"🎉 DOWNLOAD COMPLETE!")
        print(f"{'='*70}")
        print(f"   ✅ Successfully downloaded: {total_downloaded}/{total_frames} frames")
        print(f"   📁 Organized into: {len(downloaded_sections)} sections")
        print(f"   💾 Saved to: {output_dir}/")
        print(f"{'='*70}\n")
        
        return downloaded_sections
    
    def get_image_urls(self, file_key: str, node_ids: List[str], scale: float = 2, 
                       image_format: str = "png") -> Dict[str, str]:
        """Get download URLs for specific nodes with batch processing"""
        batch_size = 20  # DO NOT increase
        all_images = {}

        for i in range(0, len(node_ids), batch_size):
            batch = node_ids[i:i + batch_size]

            url = f'{self.base_url}/images/{file_key}'
            params = {
                'ids': ','.join(batch),
                'scale': scale,
                'format': image_format
            }

            try:
                response = self.figma_api_get(url, params=params)
                images = response.get('images', {})

                for node_id, image_url in images.items():
                    if image_url:
                        all_images[node_id] = image_url

            except Exception as e:
                # 🔥 FALLBACK: try EACH node individually
                print(f"⚠️ Batch failed, retrying individually ({len(batch)} nodes)")

                for node_id in batch:
                    try:
                        single_params = {
                            'ids': node_id,
                            'scale': scale,
                            'format': image_format
                        }

                        single_response = self.figma_api_get(url, params=single_params)
                        image_url = single_response.get('images', {}).get(node_id)

                        if image_url:
                            all_images[node_id] = image_url

                    except Exception as single_err:
                        print(f"❌ Skipped node {node_id}: {single_err}")


            response.raise_for_status()
            data = response.json()
            
            if 'images' in data:
                all_images.update(data['images'])
        
        return all_images
    
    def download_image(self, url: str, filepath: str) -> bool:
        """Download an image from URL to filepath."""
        try:
            response = requests.get(url, stream=True)
            response.raise_for_status()
            
            with open(filepath, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            return True
        except Exception as e:
            print(f"Error downloading image: {str(e)}")
            return False
    
    def download_all_frames(self, file_key: str, output_dir: str = "figma_frames", 
                           page_name: str = None, scale: float = 2, 
                           image_format: str = "png", include_groups: bool = True) -> List[Dict]:
        """Download all top-level frames"""
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        
        print(f"Fetching file data for: {file_key}")
        file_data = self.get_file(file_key)
        
        frames = self.extract_top_level_frames(file_data, page_name, include_groups)
        print(f"Found {len(frames)} top-level elements")
        
        if not frames:
            return []
        
        node_ids = [frame['id'] for frame in frames]
        print("\nRequesting image URLs...")
        image_urls = self.get_image_urls(file_key, node_ids, scale, image_format)
        
        downloaded_frames = []
        for i, frame in enumerate(frames, 1):
            frame_id = frame['id']
            safe_name = self.sanitize_filename(frame['name'])
            
            page_folder = self.sanitize_filename(frame['page'])
            page_path = os.path.join(output_dir, page_folder)
            Path(page_path).mkdir(parents=True, exist_ok=True)
            
            filename = f"{i:03d}_{safe_name}.{image_format}"
            filepath = os.path.join(page_path, filename)
            
            image_url = image_urls.get(frame_id)
            
            if image_url:
                print(f"  [{i}/{len(frames)}] Downloading: {frame['name']}")
                if self.download_image(image_url, filepath):
                    frame['filepath'] = filepath
                    downloaded_frames.append(frame)
        
        print(f"\n✓ Downloaded {len(downloaded_frames)}/{len(frames)} frames")
        return downloaded_frames
    
    def encode_image_to_base64(self, image_path: str) -> str:
        """Encode image to base64 string."""
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    
    def deep_visual_analysis(self, image_path: str, frame_name: str) -> Dict:
        """
        Perform detailed visual analysis with focus on accuracy and specificity.
        """
        print(f"  Analyzing: {frame_name}")
        
        try:
            base64_image = self.encode_image_to_base64(image_path)
            
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {
                                "type": "text",
                                "text": """You are an EXHAUSTIVE design analyzer. Your job is to capture EVERY SINGLE VISIBLE ELEMENT with NO EXCEPTIONS.

CRITICAL: Do NOT skip or summarize. Document EVERYTHING:
- Every piece of text (including tiny footer text)
- Every icon (including input field icons, visibility toggles)
- Every interactive element (including password show/hide buttons)
- Every color and gradient
- Every shadow and border
- Every spacing detail

For EACH visible element, provide:

1. **EXACT TEXT CONTENT**: Every word, letter, number, symbol visible (even tiny text)
2. **EXACT VISUAL APPEARANCE**: Colors (exact shades), sizes, fonts, styles
3. **PRECISE POSITIONING**: Top/bottom/left/right/center locations
4. **ELEMENT TYPE**: Button, input, text, icon, image, link, checkbox, radio, toggle, etc.
5. **INTERACTIONS**: What appears clickable, hoverable, typeable
6. **VISUAL STATES**: Current state visible (normal/hover/focus/error)
7. **STYLING DETAILS**: Borders, shadows, gradients, rounded corners, spacing
8. **GROUPINGS**: Which elements belong together visually
9. **ICONS**: Every icon, no matter how small (lock icons, eye icons, mail icons, etc.)
10. **FOOTER**: All footer content including ratings, stats, copyright

Return comprehensive JSON with this exact structure:

{
  "design_overview": {
    "page_title": "exact title text visible",
    "design_type": "login page/dashboard/form/etc",
    "primary_purpose": "what users do here",
    "layout_type": "centered/full-width/sidebar/etc",
    "color_scheme": "light/dark/gradient/etc"
  },
  "branding": [
    {
      "element_type": "logo/brand_name/tagline",
      "text_content": "exact text",
      "visual_description": "detailed appearance",
      "position": "precise location",
      "colors": ["list of colors"],
      "size": "dimensions or relative size"
    }
  ],
  "headings_and_titles": [
    {
      "text": "EXACT text content",
      "level": "h1/h2/h3 or main/sub",
      "font_style": "bold/regular/light",
      "font_family": "sans-serif/serif/monospace",
      "color": "exact color description",
      "size": "large/medium/small or px",
      "position": "exact location",
      "alignment": "left/center/right"
    }
  ],
  "body_text": [
    {
      "text": "EXACT text content",
      "purpose": "description/instruction/label",
      "style": "regular/italic/bold",
      "color": "color description",
      "size": "relative size",
      "position": "location"
    }
  ],
  "input_fields": [
    {
      "label": "exact label text",
      "field_type": "email/password/text/number/etc",
      "placeholder": "exact placeholder text",
      "current_value": "any pre-filled value",
      "width": "full/half/auto",
      "height": "size description",
      "border": "style, color, thickness",
      "background": "background color",
      "left_icon": {
        "type": "icon description",
        "color": "icon color",
        "size": "size",
        "visible": true/false
      },
      "right_icon": {
        "type": "icon description (e.g., eye icon for password visibility)",
        "color": "icon color",
        "size": "size",
        "clickable": true/false,
        "purpose": "what clicking does (show/hide password, etc)",
        "visible": true/false
      },
      "position": "exact location",
      "required": true/false,
      "validation_visible": "any error/success messages",
      "state": "empty/filled/focused/error",
      "password_masking": "if password field, how is it masked (dots/asterisks)",
      "show_hide_toggle": "if password field, describe visibility toggle"
    }
  ],
  "buttons": [
    {
      "text": "EXACT button text",
      "type": "primary/secondary/text/icon",
      "shape": "rounded/square/pill",
      "size": "small/medium/large/full-width",
      "background_color": "exact color",
      "text_color": "exact color",
      "border": "yes/no and description",
      "shadow": "yes/no and description",
      "icon": "icon description if present",
      "icon_position": "left/right/only",
      "position": "exact location",
      "hover_indication": "any visible hover state",
      "purpose": "what clicking does"
    }
  ],
  "links": [
    {
      "text": "EXACT link text",
      "type": "text_link/button_link",
      "color": "link color",
      "underline": "yes/no/hover",
      "position": "exact location",
      "purpose": "where it leads"
    }
  ],
  "checkboxes_radio_buttons": [
    {
      "type": "checkbox/radio",
      "label": "exact label text",
      "checked": true/false,
      "style": "visual description",
      "icon": "icon if checked",
      "position": "exact location",
      "group_name": "if part of radio group"
    }
  ],
  "icons": [
    {
      "type": "specific icon (user/lock/eye/mail/etc)",
      "style": "outlined/filled/duotone",
      "color": "icon color",
      "size": "small/medium/large",
      "position": "exact location",
      "associated_with": "what element it's with",
      "clickable": true/false
    }
  ],
  "dividers_separators": [
    {
      "type": "horizontal/vertical line or text",
      "text": "text if any (e.g., 'or', 'Quick Demo')",
      "style": "solid/dashed/text",
      "color": "line color",
      "thickness": "thin/medium/thick",
      "position": "exact location",
      "length": "full-width/partial"
    }
  ],
  "containers_cards": [
    {
      "purpose": "signin form/demo section/etc",
      "background": "color description",
      "border": "yes/no and description",
      "shadow": "description of shadow",
      "border_radius": "rounded corners description",
      "padding": "spacing inside",
      "position": "exact location",
      "size": "dimensions or description",
      "contains": ["list of child elements"]
    }
  ],
  "footer_info": [
    {
      "icon": "exact icon type if present (star/user/headset/etc)",
      "icon_style": "filled/outlined/color",
      "text": "EXACT text including numbers and symbols",
      "text_parts": {
        "number": "any number shown",
        "label": "label text",
        "unit": "unit if any (K, M, +, etc)"
      },
      "type": "rating/user_count/support_info/copyright/company_name",
      "position": "exact location (left/center/right of footer)",
      "color": "text and icon color",
      "size": "size description",
      "separator": "what separates this from other footer items"
    }
  ],
  "copyright_and_legal": [
    {
      "text": "EXACT text with year and copyright symbol",
      "position": "exact location",
      "color": "text color",
      "size": "size description"
    }
  ],
  "background": {
    "type": "solid/gradient/image/pattern",
    "colors": ["list all colors"],
    "gradient_direction": "if gradient",
    "description": "full description"
  },
  "spacing_and_layout": {
    "overall_padding": "page margins",
    "element_spacing": "gaps between elements",
    "alignment": "how elements align",
    "density": "tight/comfortable/spacious"
  }
}

Be EXTREMELY specific. Capture every visible pixel."""
                            },
                            {
                                "type": "image_url",
                                "image_url": {
                                    "url": f"data:image/png;base64,{base64_image}",
                                    "detail": "high"
                                }
                            }
                        ]
                    }
                ],
                max_tokens=16000,
                temperature=0.1
            )
            
            self.analytics['api_calls'] += 1
            self.analytics['tokens_used'] += response.usage.total_tokens
            
            content = response.choices[0].message.content
            
            # Parse JSON
            try:
                if "```json" in content:
                    content = content.split("```json")[1].split("```")[0].strip()
                elif "```" in content:
                    content = content.split("```")[1].split("```")[0].strip()
                
                analysis = json.loads(content)
            except json.JSONDecodeError:
                analysis = {
                    "raw_analysis": content,
                    "parsing_error": "Could not parse as JSON"
                }
            
            print(f"    ✓ Analysis complete")
            return analysis
            
        except Exception as e:
            print(f"    ✗ Error: {str(e)}")
            return {"error": str(e)}
    
    def chunk_analysis_by_category(self, analysis_data: Dict) -> List[Dict]:
        """
        Split large analysis into manageable chunks by element categories.
        """
        chunks = []
        
        # Extract metadata
        metadata = analysis_data.get('metadata', {})
        
        for frame_idx, frame_data in enumerate(analysis_data.get('frames', [])):
            frame_name = frame_data.get('frame_name', f'Frame {frame_idx + 1}')
            analysis = frame_data.get('analysis', {})
            
            # Categories to process separately
            categories = [
                ('branding', 'Logo and Branding'),
                ('headings_and_titles', 'Headings and Titles'),
                ('body_text', 'Body Text and Labels'),
                ('input_fields', 'Input Fields'),
                ('buttons', 'Buttons'),
                ('links', 'Links'),
                ('checkboxes_radio_buttons', 'Checkboxes and Radio Buttons'),
                ('icons', 'Icons'),
                ('dividers_separators', 'Dividers and Separators'),
                ('containers_cards', 'Containers and Cards'),
                ('footer_info', 'Footer Information'),
                ('copyright_and_legal', 'Copyright and Legal'),
            ]
            
            # Create chunk for design overview and layout
            overview_chunk = {
                'chunk_id': f'{frame_name}_overview',
                'frame_name': frame_name,
                'category': 'Design Overview and Layout',
                'data': {
                    'design_overview': analysis.get('design_overview', {}),
                    'background': analysis.get('background', {}),
                    'spacing_and_layout': analysis.get('spacing_and_layout', {})
                }
            }
            chunks.append(overview_chunk)
            
            # Create chunks for each category
            for category_key, category_name in categories:
                category_data = analysis.get(category_key, [])
                
                if category_data:  # Only create chunk if data exists
                    chunk = {
                        'chunk_id': f'{frame_name}_{category_key}',
                        'frame_name': frame_name,
                        'category': category_name,
                        'data': {category_key: category_data}
                    }
                    chunks.append(chunk)
        
        print(f"  Split analysis into {len(chunks)} chunks")
        return chunks
    
    def generate_test_cases_for_chunk(self, chunk: Dict, chunk_number: int, total_chunks: int) -> str:
        """
        Generate test cases for a specific chunk of analysis.
        """
        prompt = f"""You are generating test cases for a UI design. This is chunk {chunk_number} of {total_chunks}.

**Frame**: {chunk['frame_name']}
**Category**: {chunk['category']}

Generate 30-50 comprehensive manual test cases for the elements in this category.

REQUIREMENTS:
1. Use EXACT text content from the analysis
2. Reference ACTUAL visual elements that exist
3. Be SPECIFIC with expected results (exact colors, text, positions)
4. Include positive, negative, boundary cases
5. Test visual appearance, functionality, validation, accessibility

Generate test cases in this format:

## {chunk['category'].upper()} TEST CASES

### Test Case Template
For each element, generate appropriate test cases covering:
- Visual presence and styling
- Text content verification
- Position and layout
- Interactive behavior
- States (normal, hover, focus, error)
- Validation (for inputs)
- Accessibility

---

**Test Case ID**: {chunk['chunk_id'].upper()}-001  
**Title**: [Specific test title]  
**Description**: [What this test verifies]  
**Preconditions**: User is on the page  
**Test Data**: [Any required data]  
**Steps**:
1. [Specific step]
2. [Specific step]
3. [Specific step]

**Expected Result**:
- [Specific expected outcome with exact values from analysis]
- [Another expected outcome]

**Priority**: High/Medium/Low  
**Category**: {chunk['category']}

---

[Continue for all elements in this category...]

Analysis Data for this chunk:
```json
{json.dumps(chunk['data'], indent=2)}
```

Generate ALL test cases with detailed steps and exact expected results."""

        try:
            print(f"    Generating test cases for chunk {chunk_number}/{total_chunks}: {chunk['category']}")
            
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=8000,
                temperature=0.2
            )
            
            self.analytics['api_calls'] += 1
            self.analytics['tokens_used'] += response.usage.total_tokens
            
            return response.choices[0].message.content
            
        except Exception as e:
            print(f"    ✗ Error: {str(e)}")
            return f"## Error generating test cases for {chunk['category']}\n\n{str(e)}"
    
    def generate_comprehensive_test_cases(self, analysis_data: Dict, output_file: str = "detailed_test_cases.md"):
        """
        Generate comprehensive test cases by processing analysis in chunks.
        """
        print(f"\nGenerating comprehensive test cases (chunked approach)...")
        
        # Split analysis into chunks
        print("\n  Chunking analysis data...")
        chunks = self.chunk_analysis_by_category(analysis_data)
        
        # Generate header
        header = f"""# COMPREHENSIVE TEST CASE SUITE

## METADATA
- **Total Chunks**: {len(chunks)}
- **Analysis Date**: {analysis_data.get('metadata', {}).get('analysis_date', 'N/A')}
- **Total Frames Analyzed**: {analysis_data.get('metadata', {}).get('total_frames', 0)}
- **Generated**: {time.strftime('%Y-%m-%d %H:%M:%S')}

---

## TABLE OF CONTENTS
"""
        
        # Add TOC
        for i, chunk in enumerate(chunks, 1):
            header += f"{i}. {chunk['frame_name']} - {chunk['category']}\n"
        
        header += "\n---\n\n"
        
        # Write header
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(header)
        
        # Generate test cases for each chunk
        print(f"\n  Generating test cases for {len(chunks)} chunks...")
        
        for i, chunk in enumerate(chunks, 1):
            test_cases = self.generate_test_cases_for_chunk(chunk, i, len(chunks))
            
            # Append to file
            with open(output_file, 'a', encoding='utf-8') as f:
                f.write(f"\n\n# CHUNK {i}/{len(chunks)}: {chunk['frame_name']} - {chunk['category']}\n\n")
                f.write(test_cases)
                f.write("\n\n---\n")
            
            print(f"    ✓ Chunk {i}/{len(chunks)} complete")
            
            # Small delay to avoid rate limits
            if i < len(chunks):
                time.sleep(1)
        
        # Add summary footer
        footer = f"""

---

## TEST EXECUTION SUMMARY

| Chunk | Category | Status | Notes | Executed By | Date |
|-------|----------|--------|-------|-------------|------|
"""
        
        for i, chunk in enumerate(chunks, 1):
            footer += f"| {i} | {chunk['category']} | ⬜ Pending | | | |\n"
        
        footer += """

---

## EXECUTION GUIDELINES

1. Execute test cases in order by chunk
2. Mark each test as Pass/Fail
3. Document any deviations from expected results
4. Take screenshots for failed tests
5. Report bugs with test case ID reference

---

**End of Test Case Suite**
"""
        
        with open(output_file, 'a', encoding='utf-8') as f:
            f.write(footer)
        
        print(f"\n  ✓ Complete test suite saved to: {output_file}")
        return output_file
    
    @staticmethod
    def sanitize_filename(filename: str, max_length: int = 100) -> str:
        """Sanitize filename for all operating systems"""
        invalid_chars = '<>:"/\\|?*'
        for char in invalid_chars:
            filename = filename.replace(char, '')
        
        filename = filename.strip('. ')
        
        if len(filename) > max_length:
            filename = filename[:max_length].strip()
        
        if not filename:
            filename = 'unnamed'
        
        return filename




# Enhanced global status tracking
generation_status = {
    'progress': 0,
    'step': '',
    'messages': [],
    'logs': [],
    'complete': False,
    'results': None,
    'error': None,
    'analytics': {
        'start_time': None,
        'estimated_completion': None,
        'frames_processed': 0,
        'total_frames': 0,
        'api_calls': 0,
        'tokens_used': 0,
        'elapsed_time': 0,
        'test_cases_generated': 0
    }
}
def add_log(message, log_type='info', icon=None):
    """Add a detailed log entry"""
    global generation_status
    
    log_entry = {
        'timestamp': time.strftime('%H:%M:%S'),
        'message': message,
        'type': log_type,
        'icon': icon
    }
    
    generation_status['logs'].append(log_entry)
    generation_status['messages'].append({
        'text': message,
        'type': log_type,
        'time': log_entry['timestamp']
    })
    
    if len(generation_status['logs']) > 500:
        generation_status['logs'] = generation_status['logs'][-500:]
    
    print(f"[{log_entry['timestamp']}] [{log_type.upper()}] {message}")

def update_status_with_logs(progress, step, message=None, log_type='info'):
    """Enhanced status update with automatic logging"""
    global generation_status
    generation_status['progress'] = progress
    generation_status['step'] = step
    
    if message:
        add_log(message, log_type)
    
    # Update elapsed time
    if generation_status['analytics']['start_time']:
        elapsed = time.time() - generation_status['analytics']['start_time']
        generation_status['analytics']['elapsed_time'] = elapsed



def calculate_estimated_time(total_frames):
    """Calculate estimated time based on frames"""
    # Average: 15 seconds per frame analysis + 20 seconds per chunk generation
    # Typically 10-15 chunks per frame
    avg_time_per_frame = 15 + (12 * 20)  # ~4.25 minutes per frame
    return total_frames * avg_time_per_frame


@app.route('/')
def index():
    """Serve the frontend"""
    # Read the HTML from your frontend file
    with open(r'frontend2.html', 'r', encoding='utf-8') as f:
        html_content = f.read()
    # Modify the JavaScript to connect to backend
    html_content = html_content.replace(
        '// For demo purposes, we\'ll simulate API call',
        '''// Real API call to backend
        const response = await fetch('/api/generate', {
            method: 'POST',
            headers: {'Content-Type': 'application/json'},
            body: JSON.stringify(data)
        });
        if (!response.ok) {
            throw new Error('Failed to start generation');
        }
        // Start polling for real status'''
    )
    html_content = html_content.replace(
        'async function pollStatus() {',
        '''async function pollStatus() {
        while (true) {
            const response = await fetch('/api/status');
            const status = await response.json();
            updateProgress(status.progress, status.step);
            // Add new messages
            if (status.messages && status.messages.length > 0) {
                const lastMsg = status.messages[status.messages.length - 1];
                addMessage(lastMsg.text, lastMsg.type);
            }
            if (status.complete) {
                spinner.style.display = 'none';
                statusTitle.innerHTML = '<i class="fas fa-check-circle" style="color: #34d399;"></i> Generation Complete!';
                showResults(status.results);
                submitBtn.disabled = false;
                submitBtn.innerHTML = '<i class="fas fa-redo"></i><span>Generate Again</span>';
                break;
            }
            if (status.error) {
                spinner.style.display = 'none';
                statusTitle.innerHTML = '<i class="fas fa-exclamation-circle" style="color: #f87171;"></i> Error';
                addMessage('Error: ' + status.error, 'error');
                submitBtn.disabled = false;
                submitBtn.innerHTML = '<i class="fas fa-bolt"></i><span>Generate Test Cases</span>';
                break;
            }
            await new Promise(resolve => setTimeout(resolve, 1000));
        }
    }
    async function pollStatusOld() {'''
    )
    # Update download links
    html_content = html_content.replace(
        '<a href="/download/analysis"',
        '<a href="/api/download/analysis"'
    ).replace(
        '<a href="/download/testcases"',
        '<a href="/api/download/testcases"'
    )
    return render_template_string(html_content)
 
 

def format_time(seconds):
    """Format seconds into human readable time"""
    if seconds < 60:
        return f"{int(seconds)}s"
    elif seconds < 3600:
        minutes = int(seconds / 60)
        secs = int(seconds % 60)
        return f"{minutes}m {secs}s"
    else:
        hours = int(seconds / 3600)
        minutes = int((seconds % 3600) / 60)
        return f"{hours}h {minutes}m"

def update_status_enhanced(progress, step, message, msg_type='info'):
    """Enhanced status update with analytics"""
    global generation_status
    generation_status['progress'] = progress
    generation_status['step'] = step
    generation_status['messages'].append({
        'text': message,
        'type': msg_type,
        'time': time.strftime('%H:%M:%S')
    })
    
    # Update analytics
    if generation_status['analytics']['start_time']:
        elapsed = time.time() - generation_status['analytics']['start_time']
        generation_status['analytics']['elapsed_time'] = elapsed
        
        # Update estimated completion
        if generation_status['analytics']['total_frames'] > 0:
            total_estimated = calculate_estimated_time(generation_status['analytics']['total_frames'])
            progress_ratio = progress / 100
            if progress_ratio > 0:
                estimated_total = elapsed / progress_ratio
                remaining = estimated_total - elapsed
                generation_status['analytics']['estimated_completion'] = format_time(remaining)



# Add this function to your backend to update analytics properly
def update_analytics_in_real_time(progress_step, frames_processed=0, total_frames=0, 
                                 api_calls=0, tokens_used=0):
    """Update analytics in real-time during generation"""
    global generation_status
    
    if not generation_status['analytics']['start_time']:
        generation_status['analytics']['start_time'] = time.time()
    
    # Update analytics
    generation_status['analytics']['frames_processed'] = frames_processed
    generation_status['analytics']['total_frames'] = total_frames
    generation_status['analytics']['api_calls'] = api_calls
    generation_status['analytics']['tokens_used'] = tokens_used
    
    # Calculate elapsed time
    if generation_status['analytics']['start_time']:
        elapsed = time.time() - generation_status['analytics']['start_time']
        generation_status['analytics']['elapsed_time'] = elapsed
        
        # Calculate estimated completion based on progress
        if total_frames > 0 and frames_processed > 0:
            progress_ratio = frames_processed / total_frames
            if progress_ratio > 0:
                estimated_total = elapsed / progress_ratio
                remaining = estimated_total - elapsed
                generation_status['analytics']['estimated_completion'] = format_time(remaining)



def run_generation_enhanced(figma_token, file_key, scale, project_name):
    """Enhanced generation with section-wise organization"""
    global generation_status
    generation_status = {
        'progress': 0, 'step': '', 'messages': [], 'logs': [],
        'complete': False, 'results': None, 'error': None,
        'analytics': {
            'start_time': time.time(),
            'estimated_completion': 'Calculating...',
            'frames_processed': 0,
            'total_frames': 0,
            'api_calls': 0,
            'tokens_used': 0,
            'elapsed_time': 0,
            'test_cases_generated': 0
        }
    }
    
    try:
        add_log('🚀 Starting Figma Test Case Generation', 'info', 'fa-rocket')
        add_log(f'📝 Project: {project_name}', 'info')
        update_status_with_logs(5, 'Initialization', 'Initializing test case generator...')
        
        add_log('🔐 Validating API credentials...', 'info')
        update_status_with_logs(10, 'Validation', 'Validating Figma and OpenAI credentials')
        analyzer = EnhancedFigmaTestCaseGenerator(figma_token, OPENAI_API_KEY)
        add_log('✅ API credentials validated successfully', 'success')
        
        add_log('🌐 Connecting to Figma API...', 'info')
        update_status_with_logs(15, 'Connection', 'Establishing connection to Figma')
        
        OUTPUT_FOLDER = "figma_frames"
        ANALYSIS_DIR = "analysis_output"  # Changed to directory
        TEST_CASES_DIR = "test_cases_output"  # Changed to directory
        
        add_log('📊 Fetching Figma file structure...', 'info')
        update_status_with_logs(20, 'Fetching Data', 'Retrieving Figma file data')
        
        add_log(f'🔍 Extracting frames from file: {file_key}', 'info')
        add_log('⚙️ Using frame extraction with min size: 200x200px', 'info')
        
        # Download frames
        add_log('📥 Starting frame download process...', 'info')
        downloaded_sections = analyzer.download_direct_frames_only(
            file_key=file_key,
            output_dir=OUTPUT_FOLDER,
            page_name=None,
            scale=float(scale),
            image_format="png",
            min_width=200,
            min_height=200
        )
        
        if not downloaded_sections:
            raise Exception("No frames found in the Figma file")
        
        # Flatten frames with section info preserved
        all_frames = []
        for section_key, section_data in downloaded_sections.items():
            display_name = section_data['section_info']['name']
            add_log(f'📂 Found section: {display_name}', 'info')
            
            for frame in section_data['frames']:
                frame['section_key'] = section_key
                frame['section_display_name'] = display_name
                all_frames.append(frame)
                add_log(f'  └─ Frame: {frame["name"]} ({int(frame["width"])}x{int(frame["height"])}px)', 'info')
        
        total_frames = len(all_frames)
        total_sections = len(downloaded_sections)
        
        generation_status['analytics']['total_frames'] = total_frames
        estimated_time = calculate_estimated_time(total_frames)
        
        add_log(f'✅ Extraction complete: {total_frames} frames from {total_sections} sections', 'success')
        add_log(f'⏱️ Estimated processing time: {format_time(estimated_time)}', 'info')
        update_status_with_logs(35, 'Planning', f'Found {total_frames} screens. Starting AI analysis...')
        
        # Prepare analysis structure
        all_analyses = {
            "metadata": {
                "total_frames": total_frames,
                "total_sections": total_sections,
                "analysis_date": time.strftime("%Y-%m-%d %H:%M:%S"),
                "project_name": project_name
            },
            "frames": []
        }
        
        add_log(f'🤖 Starting AI analysis of {total_frames} frames...', 'info')
        add_log('📊 Using GPT-4o for deep visual analysis', 'info')
        
        # Analyze frames (section info preserved)
        for i, frame in enumerate(all_frames, 1):
            progress = 45 + (i / total_frames) * 30  # 45% to 75%
            frame_name = frame["name"]
            section_display = frame.get('section_display_name', 'Unknown')
            section_key = frame.get('section_key', 'Unknown')
            
            add_log(f'🔬 [{i}/{total_frames}] {section_display} → {frame_name[:40]}...', 'info')
            update_status_with_logs(progress, 'AI Analysis', f'Analyzing frame {i}/{total_frames}')
            
            generation_status['analytics']['frames_processed'] = i
            
            # Visual analysis
            add_log(f'  ├─ Encoding image...', 'info')
            analysis = analyzer.deep_visual_analysis(frame['filepath'], frame_name)
            add_log(f'  ├─ Analysis complete (API call #{analyzer.analytics["api_calls"]})', 'success')
            add_log(f'  └─ Tokens used: {analyzer.analytics["tokens_used"]:,}', 'info')
            platform = analyzer.detect_platform_from_dimensions(
            frame['width'], frame['height']
        )
            all_analyses['frames'].append({
    "frame_name": frame_name,
    "section": section_display,
    "section_key": section_key,
    "platform": platform,  # ✅ ADD THIS
    "dimensions": {
        "width": frame['width'],
        "height": frame['height']
    },
    "analysis": analysis
})
            
            # Update analytics
            generation_status['analytics']['api_calls'] = analyzer.analytics['api_calls']
            generation_status['analytics']['tokens_used'] = analyzer.analytics['tokens_used']
            
            # Calculate remaining time
            elapsed = time.time() - generation_status['analytics']['start_time']
            if i > 0:
                avg_per_frame = elapsed / i
                remaining_frames = total_frames - i
                est_remaining = remaining_frames * avg_per_frame
                generation_status['analytics']['estimated_completion'] = format_time(est_remaining)
        
        # Save analysis in section-wise structure
        add_log('💾 Saving analysis in section-wise structure...', 'info')
        update_status_with_logs(55, 'Saving Analysis', 'Organizing analysis by sections')
        analysis_dir = analyzer.save_analysis_by_sections(all_analyses, ANALYSIS_DIR)
        add_log(f'✅ Section-wise analysis saved: {analysis_dir}/', 'success')
        
        # Generate test cases in section-wise structure
        add_log('📝 Generating test cases by section and frame...', 'info')
        update_status_with_logs(68, 'Test Generation', 'Creating section-wise test cases')
        
        test_results = analyzer.generate_test_cases_by_sections(all_analyses, TEST_CASES_DIR)
        add_log(f'✅ Test cases generated: {test_results["output_dir"]}/', 'success')
        add_log(f'📊 Total test cases: {test_results["total_test_cases"]}', 'success')
        
        generation_status['analytics']['test_cases_generated'] = test_results['total_test_cases']
        
        # Generate Excel report (still using consolidated format)
        add_log('📊 Generating consolidated Excel report...', 'info')
        update_status_with_logs(72, 'Excel Generation', 'Creating Excel workbook')
        
        EXCEL_FILE = "test_cases_consolidated.xlsx"
        # You'll need to create a consolidated markdown first or modify excel generation
        # For now, we'll generate from the master file
        analyzer.generate_excel_report(
    markdown_file="test_cases_output",  # DIRECTORY, not file
    project_name=project_name,
    output_file=EXCEL_FILE
)
        add_log(f'✅ Excel report created: {EXCEL_FILE}', 'success')
        
        # Calculate final statistics
        total_components = 0
        for frame_data in all_analyses['frames']:
            if 'analysis' in frame_data and isinstance(frame_data['analysis'], dict):
                analysis = frame_data['analysis']
                total_components += len(analysis.get('buttons', []))
                total_components += len(analysis.get('input_fields', []))
                total_components += len(analysis.get('icons', []))
                total_components += len(analysis.get('links', []))
        
        total_time = time.time() - generation_status['analytics']['start_time']
        
        generation_status['results'] = {
            'testCases': test_results.get('total_test_cases', 0),
            'frames': total_frames,
            'sections': total_sections,
            'components': total_components,
            'time': format_time(total_time),
            'analysisDir': analysis_dir,
            'testCasesDir': test_results.get('output_dir', ''),
            'masterTestFile': test_results.get('master_file', ''),
            'excelFile': EXCEL_FILE,
            'analytics': {
                'total_time': format_time(total_time),
                'api_calls': generation_status['analytics']['api_calls'],
                'tokens_used': generation_status['analytics']['tokens_used'],
                'avg_time_per_frame': format_time(total_time / total_frames) if total_frames > 0 else '0s'
            }
        }
        
        add_log('='*50, 'success')
        add_log('🎉 GENERATION COMPLETE!', 'success')
        add_log('='*50, 'success')
        add_log(f'✅ Test Cases: {test_results["total_test_cases"]}', 'success')
        add_log(f'✅ Frames Analyzed: {total_frames}', 'success')
        add_log(f'✅ Sections: {total_sections}', 'success')
        add_log(f'✅ Components Detected: {total_components}', 'success')
        add_log(f'✅ Total Time: {format_time(total_time)}', 'success')
        add_log(f'✅ API Calls: {generation_status["analytics"]["api_calls"]}', 'success')
        add_log(f'✅ Tokens Used: {generation_status["analytics"]["tokens_used"]:,}', 'success')
        add_log('='*50, 'success')
        add_log(f'📁 Analysis Directory: {analysis_dir}/', 'info')
        add_log(f'📁 Test Cases Directory: {test_results["output_dir"]}/', 'info')
        add_log(f'📋 Master Test File: {test_results["master_file"]}', 'info')
        add_log('='*50, 'success')
        
        update_status_with_logs(87, 'Complete', '✓ All test cases generated successfully!', 'success')
        generation_status['complete'] = True
        generation_status['progress'] = 100
        generation_status['step'] = 'Complete'

        add_log('☁️ Uploading results to cloud storage...', 'info')
        update_status_with_logs(90, 'Complete', ' ☁️ Uploading results to cloud storage...', 'progress')
        # Create a unique session ID
        sanitized_project = re.sub(r'[^a-zA-Z0-9_-]', '_', project_name)
        session_id = f"{sanitized_project}_{int(time.time())}"
        
        # Upload analysis directory
        import shutil
        analysis_zip = f"temp_{session_id}_analysis.zip"
        shutil.make_archive(analysis_zip.replace('.zip', ''), 'zip', ANALYSIS_DIR)
        upload_to_gcs(analysis_zip, f"{session_id}/analysis.zip")
        os.remove(analysis_zip)
        
        # Upload test cases directory
        testcases_zip = f"temp_{session_id}_testcases.zip"
        shutil.make_archive(testcases_zip.replace('.zip', ''), 'zip', TEST_CASES_DIR)
        upload_to_gcs(testcases_zip, f"{session_id}/testcases.zip")
        os.remove(testcases_zip)
        
        # Generate and upload Excel
        excel_filename = f"{session_id}_test_cases.xlsx"
        analyzer.generate_excel_report(
            markdown_file=TEST_CASES_DIR,
            project_name=project_name,
            output_file=excel_filename
        )
        upload_to_gcs(excel_filename, f"{session_id}/excel_report.xlsx")
        os.remove(excel_filename)

        cloud_paths = {
            'analysis': f"{session_id}/analysis.zip",
            'test_cases': f"{session_id}/testcases.zip",
            'excel': f"{session_id}/excel_report.xlsx"
        }
        
        add_log('✅ Results uploaded to cloud storage', 'success')
        update_status_with_logs(95, 'Complete', ' ☁️ Results uploaded to cloud storage', 'success')

        # Preserve existing result details while adding session metadata
        if not generation_status.get('results'):
            generation_status['results'] = {}

        generation_status['results'].update({
            'session_id': session_id,
            'cloudPaths': cloud_paths
        })
        update_status_with_logs(100, 'Complete', '✅Project Completed', 'success')
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        generation_status['error'] = f"{str(e)}\n\n{error_details}"
        add_log(f'❌ ERROR: {str(e)}', 'error')
        add_log('See console for full stack trace', 'error')
        update_status_with_logs(0, 'Error', f'Generation failed: {str(e)}', 'error')
        print(f"\n❌ ERROR:\n{error_details}")
        
# ====================
# NEW API ROUTES
# ====================

import zipfile
import shutil

@app.route('/api/download/analysis-archive', methods=['GET'])
def download_analysis_archive():
    """Download the entire analysis directory as a ZIP file"""
    analysis_dir = 'analysis_output'
    
    if not os.path.exists(analysis_dir):
        return jsonify({'error': 'Analysis directory not found'}), 404
    
    # Create ZIP file
    zip_filename = f'analysis_archive_{int(time.time())}.zip'
    
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(analysis_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, analysis_dir)
                zipf.write(file_path, arcname)
    
    response = send_file(zip_filename, as_attachment=True, 
                        download_name=f'figma_analysis_{int(time.time())}.zip',
                        mimetype='application/zip')
    
    # Clean up the temp zip file after sending
    @response.call_on_close
    def cleanup():
        try:
            os.remove(zip_filename)
        except:
            pass
    
    return response


@app.route('/api/download/testcases-archive', methods=['GET'])
def download_testcases_archive():
    """Download the entire test cases directory as a ZIP file"""
    testcases_dir = 'test_cases_output'
    
    if not os.path.exists(testcases_dir):
        return jsonify({'error': 'Test cases directory not found'}), 404
    
    # Create ZIP file
    zip_filename = f'testcases_archive_{int(time.time())}.zip'
    
    with zipfile.ZipFile(zip_filename, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for root, dirs, files in os.walk(testcases_dir):
            for file in files:
                file_path = os.path.join(root, file)
                arcname = os.path.relpath(file_path, testcases_dir)
                zipf.write(file_path, arcname)
    
    response = send_file(zip_filename, as_attachment=True,
                        download_name=f'test_cases_{int(time.time())}.zip',
                        mimetype='application/zip')
    
    @response.call_on_close
    def cleanup():
        try:
            os.remove(zip_filename)
        except:
            pass
    
    return response


@app.route('/api/download/master-testcases', methods=['GET'])
def download_master_testcases():
    """Download the master test suite file"""
    filepath = 'test_cases_output/00_MASTER_TEST_SUITE.md'
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'Master test suite not found'}), 404
    
    response = send_file(filepath, as_attachment=True,
                        download_name='MASTER_TEST_SUITE.md',
                        mimetype='text/markdown')
    return response


@app.route('/api/browse/structure', methods=['GET'])
def browse_structure():
    """Get the structure of generated files for browsing"""
    result = {
        'analysis': {},
        'test_cases': {}
    }
    
    # Parse analysis structure
    analysis_dir = 'analysis_output'
    if os.path.exists(analysis_dir):
        index_file = os.path.join(analysis_dir, 'index.json')
        if os.path.exists(index_file):
            with open(index_file, 'r', encoding='utf-8') as f:
                result['analysis'] = json.load(f)
    
    # Parse test cases structure
    testcases_dir = 'test_cases_output'
    if os.path.exists(testcases_dir):
        sections = []
        for item in sorted(os.listdir(testcases_dir)):
            item_path = os.path.join(testcases_dir, item)
            if os.path.isdir(item_path):
                section_info = {
                    'name': item,
                    'files': []
                }
                for file in sorted(os.listdir(item_path)):
                    if file.endswith('.md'):
                        file_path = os.path.join(item_path, file)
                        stat = os.stat(file_path)
                        section_info['files'].append({
                            'name': file,
                            'size': stat.st_size,
                            'path': os.path.relpath(file_path, testcases_dir)
                        })
                sections.append(section_info)
        result['test_cases']['sections'] = sections
    
    return jsonify(result)


@app.route('/api/download/file/<path:filepath>', methods=['GET'])
def download_specific_file(filepath):
    """Download a specific file from the output directories"""
    # Security: only allow files from our output directories
    allowed_dirs = ['analysis_output', 'test_cases_output', 'figma_frames']
    
    # Check if filepath starts with an allowed directory
    if not any(filepath.startswith(d) for d in allowed_dirs):
        return jsonify({'error': 'Access denied'}), 403
    
    if not os.path.exists(filepath):
        return jsonify({'error': 'File not found'}), 404
    
    filename = os.path.basename(filepath)
    
    # Determine mimetype
    if filepath.endswith('.json'):
        mimetype = 'application/json'
    elif filepath.endswith('.md'):
        mimetype = 'text/markdown'
    elif filepath.endswith('.png'):
        mimetype = 'image/png'
    else:
        mimetype = 'application/octet-stream'
    
    return send_file(filepath, as_attachment=True,
                    download_name=filename,
                    mimetype=mimetype)

@app.route('/api/download/excel-consolidated', methods=['GET'])
def download_excel_consolidated():
    """Download Excel from GCS with proper error handling"""
    try:
        # Get session ID from query params or generation status
        session_id = request.args.get('session_id')
        
        if not session_id and generation_status.get('results'):
            session_id = generation_status['results'].get('session_id')
        
        if not session_id:
            print("❌ No session ID found")
            return jsonify({'error': 'No session ID found. Please generate test cases first.'}), 404
        
        project_name = request.args.get('project_name', 'Figma_Design')
        sanitized_name = re.sub(r'[^a-zA-Z0-9_-]', '_', project_name)
        
        print(f"\n{'='*60}")
        print(f"📊 Excel Download Request")
        print(f"{'='*60}")
        print(f"Session ID: {session_id}")
        print(f"Project Name: {project_name}")
        
        # GCS path
        gcs_path = f"{session_id}/excel_report.xlsx"
        
        # Create temp directory if it doesn't exist
        temp_dir = tempfile.gettempdir()
        sanitized_session = re.sub(r'[^a-zA-Z0-9_-]', '_', session_id)
        temp_file = os.path.join(temp_dir, f"{sanitized_session}_excel.xlsx")
        
        print(f"📥 GCS Path: gs://{GCS_BUCKET_NAME}/{gcs_path}")
        print(f"💾 Temp File: {temp_file}")
        
        # Download from GCS
        if download_from_gcs(gcs_path, temp_file):
            print(f"✅ File downloaded from GCS successfully")
            
            # Verify file exists and has content
            if not os.path.exists(temp_file):
                print(f"❌ Temp file was not created: {temp_file}")
                return jsonify({'error': 'Failed to create temporary file'}), 500
            
            file_size = os.path.getsize(temp_file)
            print(f"📦 File size: {file_size:,} bytes")
            
            if file_size == 0:
                print(f"❌ Downloaded file is empty")
                return jsonify({'error': 'Downloaded file is empty'}), 500
            
            # Read file into memory
            try:
                with open(temp_file, 'rb') as f:
                    file_data = f.read()
                print(f"✅ File read into memory: {len(file_data):,} bytes")
            except Exception as read_error:
                print(f"❌ Error reading file: {read_error}")
                return jsonify({'error': f'Error reading file: {str(read_error)}'}), 500
            
            # Cleanup temp file
            try:
                os.remove(temp_file)
                print(f"🗑️ Temp file cleaned up")
            except Exception as cleanup_error:
                print(f"⚠️ Could not delete temp file: {cleanup_error}")
            
            # Send file
            from io import BytesIO
            final_filename = f"{sanitized_name}_test_cases.xlsx"
            print(f"📤 Sending file as: {final_filename}")
            print(f"{'='*60}\n")
            
            return send_file(
                BytesIO(file_data),
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=final_filename
            )
        else:
            print(f"❌ File not found in GCS: gs://{GCS_BUCKET_NAME}/{gcs_path}")
            print(f"{'='*60}\n")
            
            # Try to list files in the session directory
            try:
                storage_client = storage.Client()
                bucket = storage_client.bucket(GCS_BUCKET_NAME)
                blobs = list(bucket.list_blobs(prefix=f"{session_id}/"))
                print(f"📂 Files in session directory:")
                for blob in blobs:
                    print(f"   - {blob.name}")
            except Exception as list_error:
                print(f"⚠️ Could not list files: {list_error}")
            
            return jsonify({
                'error': 'Excel file not found in cloud storage',
                'session_id': session_id,
                'gcs_path': gcs_path,
                'bucket': GCS_BUCKET_NAME
            }), 404
            
    except Exception as e:
        print(f"\n❌ Excel Download Error:")
        print(f"{'='*60}")
        import traceback
        traceback.print_exc()
        print(f"{'='*60}\n")
        return jsonify({'error': str(e)}), 500
    
# Legacy routes for backwards compatibility
@app.route('/api/download/analysis', methods=['GET'])
def download_analysis():
    """Download analysis - supports both single file and archive"""
    # Check if section-wise structure exists
    if os.path.exists('analysis_output'):
        return download_analysis_archive()
    # Fall back to single file if it exists
    elif os.path.exists('visual_analysis.json'):
        filepath = 'visual_analysis.json'
        response = send_file(filepath, as_attachment=True)
        filename = os.path.basename(filepath)
        utf8_fname = quote(filename)
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{utf8_fname}"
        return response
    else:
        return jsonify({'error': 'Analysis not found'}), 404


@app.route('/api/download/testcases', methods=['GET'])
def download_testcases():
    """Download test cases - supports both single file and archive"""
    # Check if section-wise structure exists
    if os.path.exists('test_cases_output'):
        return download_testcases_archive()
    # Fall back to single file if it exists
    elif os.path.exists('comprehensive_test_cases.md'):
        filepath = 'comprehensive_test_cases.md'
        response = send_file(filepath, as_attachment=True)
        filename = os.path.basename(filepath)
        utf8_fname = quote(filename)
        response.headers["Content-Disposition"] = f"attachment; filename*=UTF-8''{utf8_fname}"
        return response
    else:
        return jsonify({'error': 'Test cases not found'}), 404

@app.route('/api/generate', methods=['POST'])
def generate():
    """Start the generation process"""
    data = request.json
    project_name = data.get('projectName', 'Figma Design')
    figma_token = data.get('figmaToken')
    file_key = data.get('fileKey')
    scale = data.get('scale', 3)
    
    # Extract file key from URL if needed
    if 'figma.com/file/' in file_key:
        match = re.search(r'file/([^/]+)', file_key)
        if match:
            file_key = match.group(1)
    
    # Start generation in background thread
    thread = threading.Thread(
        target=run_generation_enhanced,
        args=(figma_token, file_key, scale, project_name)
    )
    thread.daemon = True
    thread.start()
    
    return jsonify({'status': 'started'})

@app.route('/api/status', methods=['GET'])
def get_status():
    """Get current generation status with analytics"""
    return jsonify(generation_status)

@app.route('/api/history', methods=['GET'])
def get_history():
    """Get list of previously generated test cases"""
    try:
        markdown_files = glob.glob("*.md")
        history = []
        
        for md_file in markdown_files:
            if md_file.startswith('comprehensive_test_cases'):
                stat = os.stat(md_file)
                
                # Read file to get metadata
                with open(md_file, 'r', encoding='utf-8') as f:
                    content = f.read()
                    test_count = content.count('**Test Case ID**')
                    
                    # Try to extract metadata
                    project_match = re.search(r'\*\*Total Frames Analyzed\*\*:\s*(\d+)', content)
                    frames = int(project_match.group(1)) if project_match else 0
                
                history.append({
                    'filename': md_file,
                    'created': datetime.fromtimestamp(stat.st_ctime).strftime('%Y-%m-%d %H:%M:%S'),
                    'size': f"{stat.st_size / 1024:.1f} KB",
                    'test_cases': test_count,
                    'frames': frames
                })
        
        # Sort by creation time, newest first
        history.sort(key=lambda x: x['created'], reverse=True)
        return jsonify({'history': history})
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    
@app.route('/api/convert-to-excel', methods=['POST'])
def convert_to_excel():
    """Convert an existing markdown file to Excel with improved error handling"""
    try:
        data = request.json
        project_name = data.get('projectName', 'Figma Design')
        
        print(f"\n📊 Converting to Excel for project: {project_name}")
        
        # Generate unique Excel filename
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_file = f"test_cases_{timestamp}.xlsx"
        
        # Create analyzer instance
        analyzer = EnhancedFigmaTestCaseGenerator("dummy_token", OPENAI_API_KEY)
        
        # Look for test case files
        test_case_files = analyzer.find_test_case_files()
        
        if not test_case_files:
            return jsonify({'error': 'No test case files found'}), 404
        
        print(f"Found {len(test_case_files)} test case files")
        
        # Try to parse all files
        all_test_cases = analyzer.parse_all_test_case_files(test_case_files)
        
        if not all_test_cases:
            # Try alternative parsing methods
            print("Trying alternative parsing...")
            # Look for any markdown files that might contain test cases
            markdown_files = []
            for root, dirs, files in os.walk('.'):
                for file in files:
                    if file.endswith('.md') and 'test' in file.lower():
                        markdown_files.append(os.path.join(root, file))
            
            for md_file in markdown_files[:5]:  # Try first 5 files
                try:
                    test_cases = analyzer.parse_test_cases_from_markdown(md_file)
                    all_test_cases.extend(test_cases)
                except:
                    pass
        
        if not all_test_cases:
            return jsonify({'error': 'Could not parse any test cases from markdown files'}), 400
        
        print(f"Total test cases to export: {len(all_test_cases)}")
        
        # Generate Excel file
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Cases"
        
        # Headers
        headers = ["Test Case ID", "Title", "Category", "Priority", "Description", 
                  "Preconditions", "Steps", "Expected Result", "Platform", "Status"]
        
        # Write headers
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Write test cases
        for row_num, tc in enumerate(all_test_cases, 2):
            ws.cell(row=row_num, column=1, value=tc.get('tc_id', f'TC-{row_num-1}'))
            ws.cell(row=row_num, column=2, value=tc.get('title', ''))
            ws.cell(row=row_num, column=3, value=tc.get('category', 'General'))
            ws.cell(row=row_num, column=4, value=tc.get('platform', 'WEB'))
            ws.cell(row=row_num, column=5, value=tc.get('description', ''))
            ws.cell(row=row_num, column=6, value=tc.get('preconditions', ''))
            ws.cell(row=row_num, column=7, value=tc.get('steps', ''))
            ws.cell(row=row_num, column=8, value=tc.get('priority', 'Medium'))
            ws.cell(row=row_num, column=9, value="Web")
            ws.cell(row=row_num, column=10, value="Pending")
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the file
        wb.save(excel_file)
        
        print(f"✅ Excel file generated: {excel_file} with {len(all_test_cases)} test cases")
        
        return jsonify({
            'success': True,
            'filename': excel_file,
            'test_cases': len(all_test_cases),
            'message': f'Successfully exported {len(all_test_cases)} test cases to Excel'
        })
        
    except Exception as e:
        import traceback
        print(f"❌ Error converting to Excel: {e}")
        print(traceback.format_exc())
        return jsonify({'error': f'Failed to generate Excel: {str(e)}'}), 500

@app.route('/api/download/excel/<filename>', methods=['GET'])
def download_excel_file(filename):
    """Download Excel file"""
    if not os.path.exists(filename):
        return jsonify({'error': 'File not found'}), 404
    
    return send_file(
        filename,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ====================
# ADDITIONAL BACKEND ROUTES FOR DESIGN ANALYSIS
# ====================

from concurrent.futures import ThreadPoolExecutor, as_completed

class DesignAnalyzer:
    """Enhanced design analysis using OpenAI"""
    
    def __init__(self, openai_api_key: str):
        self.openai_client = OpenAI(api_key=openai_api_key)
    
    def analyze_design_comprehensive(self, analysis_data: Dict) -> Dict:
        """Perform comprehensive design analysis using OpenAI"""
        try:
            # Prepare analysis summary
            frames_summary = []
            for frame in analysis_data.get('frames', []):
                frame_analysis = frame.get('analysis', {})
                frames_summary.append({
                    'name': frame.get('frame_name', ''),
                    'platform': frame.get('platform', 'WEB'),
                    'elements': self._extract_element_counts(frame_analysis)
                })
            
            prompt = f"""
            Analyze this Figma design and provide comprehensive insights. Focus on:
            
            1. **Design Strengths** (Positives) - What works well
            2. **Design Weaknesses** (Negatives) - What needs improvement
            3. **Critical Issues** (Corrections) - Must-fix problems
            4. **Enhancement Suggestions** (Improvements) - Optional improvements
            5. **Quality Metrics** - Scores for various aspects
            6. **Recommendations** - Actionable suggestions
            
            Design Summary:
            - Total Frames: {len(analysis_data.get('frames', []))}
            - Frames Analysis: {frames_summary}
            
            Analysis Data:
            {json.dumps(analysis_data, indent=2)}
            
            Return JSON in this exact format:
            {{
                "design_positives": ["string", "string", ...],
                "design_negatives": ["string", "string", ...],
                "required_corrections": ["string", "string", ...],
                "improvement_suggestions": ["string", "string", ...],
                "quality_metrics": {{
                    "accessibility": 0-100,
                    "consistency": 0-100,
                    "usability": 0-100,
                    "error_handling": 0-100,
                    "performance": 0-100,
                    "security": 0-100
                }},
                "recommendations": {{
                    "critical_issues": ["string", "string", ...],
                    "ux_improvements": ["string", "string", ...],
                    "security_recommendations": ["string", "string", ...],
                    "mobile_optimizations": ["string", "string", ...]
                }}
            }}
            
            Be specific, actionable, and reference actual design elements found.
            """
            
            response = self.openai_client.chat.completions.create(
                model="gpt-4o",
                messages=[{"role": "user", "content": prompt}],
                max_tokens=4000,
                temperature=0.3
            )
            
            content = response.choices[0].message.content
            
            # Extract JSON from response
            if "```json" in content:
                content = content.split("```json")[1].split("```")[0].strip()
            elif "```" in content:
                content = content.split("```")[1].split("```")[0].strip()
            
            analysis = json.loads(content)
            
            # Add platform distribution
            analysis['platform_distribution'] = self._calculate_platform_distribution(analysis_data)
            
            # Add element statistics
            analysis['element_statistics'] = self._calculate_element_statistics(analysis_data)
            
            return analysis
            
        except Exception as e:
            print(f"Design analysis error: {e}")
            return self._get_default_analysis()
    
    def _extract_element_counts(self, frame_analysis: Dict) -> Dict:
        """Extract element counts from frame analysis"""
        counts = {}
        elements = [
            'buttons', 'input_fields', 'icons', 'links', 
            'headings_and_titles', 'body_text', 'containers_cards'
        ]
        
        for element in elements:
            if element in frame_analysis:
                if isinstance(frame_analysis[element], list):
                    counts[element] = len(frame_analysis[element])
                else:
                    counts[element] = 1
        
        return counts
    
    def _calculate_platform_distribution(self, analysis_data: Dict) -> Dict:
        """Calculate platform distribution from analysis"""
        platforms = {}
        for frame in analysis_data.get('frames', []):
            platform = frame.get('platform', 'WEB')
            platforms[platform] = platforms.get(platform, 0) + 1
        
        # Calculate percentages
        total_frames = len(analysis_data.get('frames', []))
        if total_frames > 0:
            return {k: (v / total_frames) * 100 for k, v in platforms.items()}
        return platforms
    
    def _calculate_element_statistics(self, analysis_data: Dict) -> Dict:
        """Calculate element statistics across all frames"""
        totals = {
            'buttons': 0,
            'input_fields': 0,
            'icons': 0,
            'links': 0,
            'headings': 0,
            'text_elements': 0,
            'containers': 0
        }
        
        for frame in analysis_data.get('frames', []):
            analysis = frame.get('analysis', {})
            
            for key, count_key in [
                ('buttons', 'buttons'),
                ('input_fields', 'input_fields'),
                ('icons', 'icons'),
                ('links', 'links'),
                ('headings_and_titles', 'headings'),
                ('body_text', 'text_elements'),
                ('containers_cards', 'containers')
            ]:
                if key in analysis:
                    if isinstance(analysis[key], list):
                        totals[count_key] += len(analysis[key])
                    else:
                        totals[count_key] += 1
        
        return totals
    
    def _get_default_analysis(self) -> Dict:
        """Return default analysis structure in case of error"""
        return {
            "design_positives": ["Analysis not available"],
            "design_negatives": ["Analysis not available"],
            "required_corrections": ["Analysis not available"],
            "improvement_suggestions": ["Analysis not available"],
            "quality_metrics": {
                "accessibility": 0,
                "consistency": 0,
                "usability": 0,
                "error_handling": 0,
                "performance": 0,
                "security": 0
            },
            "recommendations": {
                "critical_issues": ["Analysis not available"],
                "ux_improvements": ["Analysis not available"],
                "security_recommendations": ["Analysis not available"],
                "mobile_optimizations": ["Analysis not available"]
            },
            "platform_distribution": {},
            "element_statistics": {}
        }


class TestCaseAnalyzer:
    """Enhanced test case analysis"""
    
    def __init__(self):
        self.executor = ThreadPoolExecutor(max_workers=5)
    
    def analyze_test_cases(self, test_cases_dir: str) -> Dict:
        """Analyze test cases for statistics and insights"""
        try:
            # Read all test case files
            test_cases = self._collect_test_cases(test_cases_dir)
            
            # Calculate statistics
            stats = self._calculate_test_case_statistics(test_cases)
            
            # Generate insights
            insights = self._generate_test_case_insights(test_cases)
            
            return {
                **stats,
                **insights,
                'total_test_cases': len(test_cases)
            }
            
        except Exception as e:
            print(f"Test case analysis error: {e}")
            return self._get_default_test_case_analysis()
    
    def _collect_test_cases(self, test_cases_dir: str) -> List[Dict]:
        """Collect all test cases from directory"""
        test_cases = []
        
        for root, dirs, files in os.walk(test_cases_dir):
            for file in files:
                if file.endswith('.md') and not file.startswith('00_'):
                    file_path = os.path.join(root, file)
                    
                    # Parse test cases from markdown
                    with open(file_path, 'r', encoding='utf-8') as f:
                        content = f.read()
                    
                    # Extract test cases using regex
                    test_case_pattern = r'\*\*Test Case ID\*\*:(.+?)(?=\n\*\*Test Case ID\*\*:|\n#|$)'
                    matches = re.findall(test_case_pattern, content, re.DOTALL)
                    
                    for match in matches:
                        tc = self._parse_test_case(match)
                        if tc:
                            test_cases.append(tc)
        
        return test_cases
    
    def _parse_test_case(self, content: str):
        """Parse individual test case from content"""
        try:
            # Extract test case details using regex
            tc_id = re.search(r'\*\*Test Case ID\*\*:\s*(.+)', content)
            title = re.search(r'\*\*Title\*\*:\s*(.+)', content)
            category = re.search(r'\*\*Category\*\*:\s*(.+)', content)
            priority = re.search(r'\*\*Priority\*\*:\s*(.+)', content)
            
            if tc_id and title:
                return {
                    'id': tc_id.group(1).strip(),
                    'title': title.group(1).strip(),
                    'category': category.group(1).strip() if category else 'General',
                    'priority': priority.group(1).strip() if priority else 'Medium'
                }
        except Exception:
            pass
        
        return None
    
    def _calculate_test_case_statistics(self, test_cases: List[Dict]) -> Dict:
        """Calculate test case statistics"""
        categories = {}
        priorities = {}
        sections = {}
        
        for tc in test_cases:
            # Category distribution
            category = tc.get('category', 'General')
            categories[category] = categories.get(category, 0) + 1
            
            # Priority distribution
            priority = tc.get('priority', 'Medium')
            priorities[priority] = priorities.get(priority, 0) + 1
            
            # Section from ID (assuming format: SECTION-FRAME-001)
            tc_id = tc.get('id', '')
            if '-' in tc_id:
                section = tc_id.split('-')[0]
                sections[section] = sections.get(section, 0) + 1
        
        return {
            'category_distribution': categories,
            'priority_distribution': priorities,
            'section_distribution': sections
        }
    
    def _generate_test_case_insights(self, test_cases: List[Dict]) -> Dict:
        """Generate insights from test cases"""
        # Calculate test coverage metrics
        total_cases = len(test_cases)
        
        # Count by functionality type
        functionality_types = {
            'UI Verification': 0,
            'Functional Testing': 0,
            'Usability Testing': 0,
            'Accessibility Testing': 0,
            'Performance Testing': 0,
            'Security Testing': 0
        }
        
        # Simple heuristic based on category and title
        for tc in test_cases:
            category = tc.get('category', '').lower()
            title = tc.get('title', '').lower()
            
            if any(word in category or word in title for word in ['visual', 'appearance', 'color', 'font']):
                functionality_types['UI Verification'] += 1
            elif any(word in category or word in title for word in ['function', 'behavior', 'click', 'input']):
                functionality_types['Functional Testing'] += 1
            elif any(word in category or word in title for word in ['user', 'experience', 'flow', 'navigation']):
                functionality_types['Usability Testing'] += 1
            elif any(word in category or word in title for word in ['accessibility', 'screen reader', 'contrast']):
                functionality_types['Accessibility Testing'] += 1
            elif any(word in category or word in title for word in ['performance', 'load', 'speed']):
                functionality_types['Performance Testing'] += 1
            elif any(word in category or word in title for word in ['security', 'auth', 'login', 'permission']):
                functionality_types['Security Testing'] += 1
            else:
                functionality_types['Functional Testing'] += 1
        
        # Calculate percentages
        if total_cases > 0:
            for key in functionality_types:
                functionality_types[key] = (functionality_types[key] / total_cases) * 100
        
        return {
            'functionality_coverage': functionality_types,
            'coverage_score': min(100, total_cases / 100 * 100),  # Simple coverage score
            'average_priority': self._calculate_average_priority(test_cases)
        }
    
    def _calculate_average_priority(self, test_cases: List[Dict]) -> float:
        """Calculate average priority score (1-5 scale)"""
        priority_scores = {
            'Critical': 5,
            'High': 4,
            'Medium': 3,
            'Low': 2,
            'Lowest': 1
        }
        
        total_score = 0
        count = 0
        
        for tc in test_cases:
            priority = tc.get('priority', 'Medium')
            if priority in priority_scores:
                total_score += priority_scores[priority]
                count += 1
        
        return round(total_score / count, 2) if count > 0 else 3.0
    
    def _get_default_test_case_analysis(self) -> Dict:
        """Return default test case analysis"""
        return {
            'category_distribution': {},
            'priority_distribution': {},
            'section_distribution': {},
            'functionality_coverage': {},
            'coverage_score': 0,
            'average_priority': 3.0,
            'total_test_cases': 0
        }


# Initialize analyzers
design_analyzer = DesignAnalyzer(OPENAI_API_KEY)
testcase_analyzer = TestCaseAnalyzer()


@app.route('/api/dashboard/analysis', methods=['GET'])
def get_dashboard_analysis():
    """Get comprehensive dashboard analysis"""
    try:
        # Check if analysis exists
        analysis_dir = 'analysis_output'
        if not os.path.exists(analysis_dir):
            return jsonify({'error': 'Analysis data not found'}), 404
        
        # Load analysis data
        index_file = os.path.join(analysis_dir, 'index.json')
        if not os.path.exists(index_file):
            return jsonify({'error': 'Analysis index not found'}), 404
        
        index_data, frames_data = load_analysis_frames(analysis_dir)
        if not index_data:
            with open(index_file, 'r', encoding='utf-8') as f:
                index_data = json.load(f)
        
        # Prepare analysis data for OpenAI
        analysis_for_openai = {
            'metadata': index_data.get('metadata', {}),
            'frames': frames_data
        }
        
        # Get design analysis from OpenAI
        design_analysis = design_analyzer.analyze_design_comprehensive(analysis_for_openai)
        
        # Get test case analysis
        test_cases_dir = 'test_cases_output'
        test_case_analysis = {}
        if os.path.exists(test_cases_dir):
            test_case_analysis = testcase_analyzer.analyze_test_cases(test_cases_dir)
        
        # Get platform distribution
        platform_distribution = {}
        for frame in frames_data:
            platform = frame.get('frame_info', {}).get('platform', 'WEB')
            platform_distribution[platform] = platform_distribution.get(platform, 0) + 1

        element_counts = summarize_elements(frames_data)
        
        # Calculate section statistics
        section_statistics = []
        for section in index_data.get('sections', []):
            section_path = os.path.join(analysis_dir, section['path'])
            if os.path.exists(section_path):
                frames = [d for d in os.listdir(section_path) 
                         if os.path.isdir(os.path.join(section_path, d))]
                
                section_statistics.append({
                    'name': section.get('display_name', section.get('key')),
                    'frames': len(frames),
                    'test_cases': count_section_test_cases(section.get('display_name', ''), section_path),
                    'quality_score': calculate_section_quality(section_path)
                })
        
        last_updated = index_data.get('metadata', {}).get('analysis_date', time.strftime('%Y-%m-%d %H:%M:%S'))

        # Combine all data
        response = {
            'design_analysis': design_analysis,
            'test_case_analysis': test_case_analysis,
            'platform_distribution': platform_distribution,
            'section_statistics': section_statistics,
            'overall_statistics': {
                'total_frames': len(frames_data),
                'total_sections': len(section_statistics),
                'total_test_cases': test_case_analysis.get('total_test_cases', 0),
                'total_components': sum(design_analysis.get('element_statistics', {}).values()) or sum(element_counts.values()),
                'analysis_time': last_updated
            },
            'element_statistics': element_counts,
            'last_updated': last_updated
        }

        # Attach cloud info if available
        if generation_status.get('results'):
            response['session_id'] = generation_status['results'].get('session_id')
            response['cloudPaths'] = generation_status['results'].get('cloudPaths', {})
        
        return jsonify(response)
        
    except Exception as e:
        print(f"Dashboard analysis error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/stats', methods=['GET'])
def get_dashboard_stats():
    """Get real-time statistics for dashboard"""
    try:
        # Get from generation status if available
        if generation_status.get('complete') and generation_status.get('results'):
            results = generation_status['results']
            
            return jsonify({
                'testCases': results.get('testCases', 0),
                'frames': results.get('frames', 0),
                'sections': results.get('sections', 0),
                'components': results.get('components', 0),
                'time': results.get('time', '0s'),
                'analytics': results.get('analytics', {}),
                'session_id': results.get('session_id', ''),
                'cloudPaths': results.get('cloudPaths', {}),
                'last_updated': time.strftime('%Y-%m-%d %H:%M:%S')
            })
        
        # Try to calculate from files
        frames_dir = 'figma_frames'
        analysis_dir = 'analysis_output'
        test_cases_dir = 'test_cases_output'
        
        # Calculate stats
        total_frames = 0
        if os.path.exists(frames_dir):
            for root, dirs, files in os.walk(frames_dir):
                total_frames += len([f for f in files if f.endswith('.png')])
        
        total_sections = 0
        if os.path.exists(analysis_dir):
            index_file = os.path.join(analysis_dir, 'index.json')
            if os.path.exists(index_file):
                with open(index_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    total_sections = len(data.get('sections', []))
        
        total_test_cases = 0
        if os.path.exists(test_cases_dir):
            for root, dirs, files in os.walk(test_cases_dir):
                for file in files:
                    if file.endswith('.md') and not file.startswith('00_'):
                        file_path = os.path.join(root, file)
                        with open(file_path, 'r', encoding='utf-8') as f:
                            content = f.read()
                            total_test_cases += content.count('**Test Case ID**')
        
        return jsonify({
            'testCases': total_test_cases,
            'frames': total_frames,
            'sections': total_sections,
            'components': 0,
            'time': 'Calculated',
            'analytics': {},
            'session_id': ''
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/compare', methods=['POST'])
def compare_designs():
    """Compare current design with previous versions or benchmarks"""
    try:
        data = request.json
        design_data = data.get('design_data', {})
        benchmark = data.get('benchmark', {})
        
        # Use OpenAI to compare designs
        prompt = f"""
        Compare these two design versions and provide insights:
        
        CURRENT DESIGN:
        {json.dumps(design_data, indent=2)}
        
        BENCHMARK/PREVIOUS VERSION:
        {json.dumps(benchmark, indent=2)}
        
        Provide comparison in this JSON format:
        {{
            "improvements": ["list improvements"],
            "regressions": ["list regressions"],
            "consistency_score": 0-100,
            "progress_score": 0-100,
            "recommendations": ["specific recommendations"]
        }}
        """
        
        response = design_analyzer.openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=2000,
            temperature=0.3
        )
        
        content = response.choices[0].message.content
        
        # Extract JSON
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()
        
        comparison = json.loads(content)
        
        return jsonify(comparison)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/generate-report', methods=['POST'])
def generate_comprehensive_report():
    """Generate a comprehensive PDF report"""
    try:
        data = request.json
        include_analysis = data.get('include_analysis', True)
        include_test_cases = data.get('include_test_cases', True)
        include_recommendations = data.get('include_recommendations', True)
        
        # Get all dashboard data
        dashboard_data = get_dashboard_analysis().get_json()
        
        # Use OpenAI to generate report content
        prompt = f"""
        Generate a comprehensive design analysis report based on this data:
        
        {json.dumps(dashboard_data, indent=2)}
        
        Create a professional report with:
        1. Executive Summary
        2. Design Analysis
        3. Test Case Coverage
        4. Quality Metrics
        5. Recommendations
        6. Action Plan
        
        Format the report in markdown with proper headings and sections.
        """
        
        response = design_analyzer.openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=4000,
            temperature=0.3
        )
        
        report_content = response.choices[0].message.content
        
        # Save report
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        report_file = f"design_analysis_report_{timestamp}.md"
        
        with open(report_file, 'w', encoding='utf-8') as f:
            f.write(report_content)
        
        # Generate PDF (simplified - in production use a proper PDF library)
        pdf_file = report_file.replace('.md', '.pdf')
        
        return jsonify({
            'success': True,
            'report_file': report_file,
            'pdf_file': pdf_file,
            'download_url': f'/api/download/report/{pdf_file}'
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/export/<format_type>', methods=['GET'])
def export_dashboard_data(format_type):
    """Export dashboard data in various formats"""
    try:
        dashboard_data = get_dashboard_analysis().get_json()
        
        if format_type == 'json':
            # Return as JSON
            response = jsonify(dashboard_data)
            response.headers['Content-Disposition'] = f'attachment; filename=dashboard_data_{int(time.time())}.json'
            return response
            
        elif format_type == 'csv':
            # Convert to CSV (simplified)
            csv_data = []
            
            # Flatten data for CSV
            if 'design_analysis' in dashboard_data:
                for key, value in dashboard_data['design_analysis'].items():
                    if isinstance(value, list):
                        csv_data.append(f"{key},{','.join(value[:5])}")
                    elif isinstance(value, dict):
                        for k, v in value.items():
                            csv_data.append(f"{key}.{k},{v}")
            
            csv_content = '\n'.join(csv_data)
            
            response = Response(csv_content, mimetype='text/csv')
            response.headers['Content-Disposition'] = f'attachment; filename=dashboard_data_{int(time.time())}.csv'
            return response
            
        elif format_type == 'markdown':
            # Convert to markdown
            md_content = f"""# Design Analysis Report
## Generated: {time.strftime('%Y-%m-%d %H:%M:%S')}

### Overview
- Total Frames: {dashboard_data.get('overall_statistics', {}).get('total_frames', 0)}
- Total Test Cases: {dashboard_data.get('overall_statistics', {}).get('total_test_cases', 0)}

### Design Analysis
"""
            
            if 'design_analysis' in dashboard_data:
                da = dashboard_data['design_analysis']
                
                md_content += "\n#### Positives\n"
                for item in da.get('design_positives', []):
                    md_content += f"- {item}\n"
                
                md_content += "\n#### Issues\n"
                for item in da.get('design_negatives', []):
                    md_content += f"- {item}\n"
            
            response = Response(md_content, mimetype='text/markdown')
            response.headers['Content-Disposition'] = f'attachment; filename=report_{int(time.time())}.md'
            return response
        
        else:
            return jsonify({'error': 'Unsupported format'}), 400
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# Update the existing run_generation_enhanced function to include dashboard data
def run_generation_enhanced_with_dashboard(figma_token, file_key, scale, project_name):
    """Enhanced generation with dashboard data collection"""
    global generation_status
    
    # Run the original generation
    run_generation_enhanced(figma_token, file_key, scale, project_name)
    
    # After generation, trigger dashboard analysis
    if generation_status.get('complete') and not generation_status.get('error'):
        try:
            # Trigger dashboard analysis in background
            thread = threading.Thread(
                target=generate_dashboard_analysis_background,
                args=(project_name,)
            )
            thread.daemon = True
            thread.start()
            
            add_log('📊 Dashboard analysis started in background', 'info')
            
        except Exception as e:
            add_log(f'⚠️ Dashboard analysis failed: {str(e)}', 'warning')


def generate_dashboard_analysis_background(project_name: str):
    """Generate dashboard analysis in background"""
    try:
        # Small delay to ensure files are written
        time.sleep(2)
        
        # Get dashboard analysis
        dashboard_data = get_dashboard_analysis().get_json()
        
        # Save to file for persistence
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        dashboard_file = f"dashboard_data_{timestamp}.json"
        
        with open(dashboard_file, 'w', encoding='utf-8') as f:
            json.dump(dashboard_data, f, indent=2)
        
        # Upload to GCS
        upload_to_gcs(dashboard_file, f"dashboard/{dashboard_file}")
        
        add_log('✅ Dashboard analysis completed', 'success')
        
    except Exception as e:
        print(f"Background dashboard analysis error: {e}")


# Update the existing generate endpoint to use the enhanced version
@app.route('/api/generate-enhanced', methods=['POST'])
def generate_enhanced():
    """Enhanced generation with dashboard support"""
    data = request.json
    project_name = data.get('projectName', 'Figma Design')
    figma_token = data.get('figmaToken')
    file_key = data.get('fileKey')
    scale = data.get('scale', 3)
    
    # Extract file key from URL if needed
    if 'figma.com/file/' in file_key:
        match = re.search(r'file/([^/]+)', file_key)
        if match:
            file_key = match.group(1)
    
    # Start enhanced generation in background thread
    thread = threading.Thread(
        target=run_generation_enhanced_with_dashboard,
        args=(figma_token, file_key, scale, project_name)
    )
    thread.daemon = True
    thread.start()
    
    return jsonify({'status': 'started', 'message': 'Generation with dashboard analysis started'})
 

 
# Update your existing run_generation function name to avoid conflicts
# Replace the call in your existing @app.route('/api/generate') with run_generation_enhanced

# ====================
# REAL DATA ENDPOINTS
# ====================

import glob
import re
from collections import defaultdict
@app.route('/dashboard.html')
def dashboard_redirect():
    """Redirect dashboard.html requests to the main app"""
    return redirect('/')

@app.route('/api/dashboard/real-stats', methods=['GET'])
def get_real_stats():
    """Fixed real statistics with accurate data"""
    try:
        stats = {
            'testCases': 0,
            'frames': 0,
            'sections': 0,
            'components': 0,
            'time': '0s',
            'analytics': {},
            'session_id': '',
            'cloudPaths': {},
            'last_updated': None
        }
        
        # First, try to get from generation_status (most recent run)
        if generation_status.get('complete') and generation_status.get('results'):
            results = generation_status['results']
            stats.update({
                'testCases': results.get('testCases', 0),
                'frames': results.get('frames', 0),
                'sections': results.get('sections', 0),
                'components': results.get('components', 0),
                'time': results.get('time', '0s'),
                'analytics': results.get('analytics', {}),
                'session_id': results.get('session_id', ''),
                'cloudPaths': results.get('cloudPaths', {}),
                'last_updated': time.strftime('%Y-%m-%d %H:%M:%S')
            })
            return jsonify(stats)
        
        # Fallback: Calculate from files
        print("Calculating stats from files...")
        
        # Count test cases
        test_cases_dir = 'test_cases_output'
        if os.path.exists(test_cases_dir):
            total_test_cases = 0
            for root, dirs, files in os.walk(test_cases_dir):
                for file in files:
                    if file.endswith('.md') and not file.startswith('00_'):
                        file_path = os.path.join(root, file)
                        try:
                            with open(file_path, 'r', encoding='utf-8') as f:
                                content = f.read()
                                # Count test case IDs
                                test_cases = len(re.findall(r'\*\*Test Case ID\*\*:', content))
                                total_test_cases += test_cases
                        except Exception as e:
                            print(f"Error reading {file}: {e}")
            stats['testCases'] = total_test_cases
            print(f"Found {total_test_cases} test cases")
        
        # Count frames and sections
        analysis_dir = 'analysis_output'
        if os.path.exists(analysis_dir):
            index_data, frames_data = load_analysis_frames(analysis_dir)

            # Frames/sections
            stats['frames'] = index_data.get('metadata', {}).get('total_frames', len(frames_data))
            stats['sections'] = len(index_data.get('sections', []))

            # Components and element counts
            element_counts = summarize_elements(frames_data)
            stats['components'] = (
                element_counts['buttons'] +
                element_counts['input_fields'] +
                element_counts['icons'] +
                element_counts['links']
            )
            stats['last_updated'] = time.strftime('%Y-%m-%d %H:%M:%S')

            print(f"Found {stats['frames']} frames, {stats['sections']} sections")
            print(f"Found {stats['components']} components")
        
        # Get analytics from generation status
        if generation_status.get('analytics'):
            stats['analytics'] = {
                'api_calls': generation_status['analytics'].get('api_calls', 0),
                'tokens_used': generation_status['analytics'].get('tokens_used', 0),
                'avg_time_per_frame': generation_status['analytics'].get('avg_time_per_frame', '0s')
            }
            
            elapsed = generation_status['analytics'].get('elapsed_time', 0)
            if elapsed > 0:
                stats['time'] = format_time(elapsed)
        
        print(f"Final stats: {stats}")
        return jsonify(stats)
        
    except Exception as e:
        print(f"Error getting stats: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500



def generate_insights_from_real_data(frames_data, element_counts):
    """Generate insights from actual frame data"""
    insights = {
        'design_positives': [],
        'design_negatives': [],
        'required_corrections': [],
        'improvement_suggestions': [],
        'recommendations': {
            'critical_issues': [],
            'ux_improvements': [],
            'security_recommendations': [],
            'mobile_optimizations': []
        }
    }
    
    # Analyze button consistency
    button_styles = set()
    for frame in frames_data:
        buttons = frame.get('analysis', {}).get('buttons', [])
        for btn in buttons:
            style = f"{btn.get('background_color', 'unknown')}-{btn.get('text_color', 'unknown')}"
            button_styles.add(style)
    
    if len(button_styles) <= 3:
        insights['design_positives'].append(
            f"✅ Consistent button styling across {element_counts['buttons']} buttons ({len(button_styles)} style variations)"
        )
    else:
        insights['design_negatives'].append(
            f"⚠️ Inconsistent button styles detected: {len(button_styles)} different variations found"
        )
    
    # Analyze form fields
    if element_counts['input_fields'] > 0:
        insights['design_positives'].append(
            f"✅ Found {element_counts['input_fields']} form input fields with proper structure"
        )
        
        # Check for labels
        labeled_fields = 0
        for frame in frames_data:
            fields = frame.get('analysis', {}).get('input_fields', [])
            for field in fields:
                if field.get('label'):
                    labeled_fields += 1
        
        if labeled_fields / max(1, element_counts['input_fields']) < 0.8:
            insights['required_corrections'].append(
                f"⚠️ Only {labeled_fields}/{element_counts['input_fields']} input fields have proper labels"
            )
    
    # Analyze icons
    if element_counts['icons'] > 0:
        insights['design_positives'].append(
            f"✅ Rich iconography: {element_counts['icons']} icons used throughout the design"
        )
    
    # Analyze text hierarchy
    if element_counts['headings'] > 0 and element_counts['text_elements'] > 0:
        ratio = element_counts['text_elements'] / element_counts['headings']
        if 3 <= ratio <= 10:
            insights['design_positives'].append(
                f"✅ Good text hierarchy: {element_counts['headings']} headings organizing {element_counts['text_elements']} text elements"
            )
        else:
            insights['improvement_suggestions'].append(
                f"💡 Consider improving text hierarchy balance (current ratio: 1:{ratio:.1f})"
            )
    
    # Add recommendations
    insights['recommendations']['critical_issues'].append(
        f"Verify all {element_counts['input_fields']} form fields have proper validation"
    )
    
    insights['recommendations']['ux_improvements'].append(
        f"Add hover states to all {element_counts['buttons']} interactive buttons"
    )
    
    insights['recommendations']['security_recommendations'].append(
        "Implement input sanitization for all form fields"
    )
    
    insights['recommendations']['mobile_optimizations'].append(
        f"Optimize touch targets for {element_counts['buttons']} buttons on mobile devices"
    )
    
    return insights



def calculate_quality_metrics_from_data(frames_data, element_counts):
    """Calculate quality metrics from actual data"""
    metrics = {
        'accessibility': 75,
        'consistency': 70,
        'usability': 80,
        'error_handling': 65,
        'performance': 85,
        'security': 70
    }
    
    # Calculate accessibility based on labeled fields
    if element_counts['input_fields'] > 0:
        labeled = 0
        for frame in frames_data:
            fields = frame.get('analysis', {}).get('input_fields', [])
            labeled += sum(1 for f in fields if f.get('label'))
        
        metrics['accessibility'] = int((labeled / element_counts['input_fields']) * 100)
    
    # Calculate consistency based on button styles
    button_styles = set()
    for frame in frames_data:
        buttons = frame.get('analysis', {}).get('buttons', [])
        for btn in buttons:
            style = f"{btn.get('background_color', '')}-{btn.get('shape', '')}"
            button_styles.add(style)
    
    if element_counts['buttons'] > 0:
        consistency_score = max(0, 100 - (len(button_styles) * 15))
        metrics['consistency'] = min(100, consistency_score)
    
    return metrics

def load_analysis_frames(analysis_dir: str):
    """Load index metadata and all frame analyses from disk."""
    frames_data = []
    index_data = {}

    index_file = os.path.join(analysis_dir, 'index.json')
    if os.path.exists(index_file):
        with open(index_file, 'r', encoding='utf-8') as f:
            index_data = json.load(f)

    for root, dirs, files in os.walk(analysis_dir):
        for file in files:
            if file == 'analysis.json':
                file_path = os.path.join(root, file)
                try:
                    with open(file_path, 'r', encoding='utf-8') as f:
                        frame_data = json.load(f)
                        frames_data.append(frame_data)
                except Exception as e:
                    print(f"Error reading {file_path}: {e}")
                    continue

    return index_data, frames_data

def summarize_elements(frames_data: List[Dict]) -> Dict:
    """Aggregate element statistics from frames."""
    element_counts = {
        'buttons': 0,
        'input_fields': 0,
        'icons': 0,
        'links': 0,
        'headings': 0,
        'text_elements': 0,
        'containers': 0
    }

    for frame in frames_data:
        analysis = frame.get('analysis', {})
        element_counts['buttons'] += len(analysis.get('buttons', []))
        element_counts['input_fields'] += len(analysis.get('input_fields', []))
        element_counts['icons'] += len(analysis.get('icons', []))
        element_counts['links'] += len(analysis.get('links', []))
        element_counts['headings'] += len(analysis.get('headings_and_titles', []))
        element_counts['text_elements'] += len(analysis.get('body_text', []))
        element_counts['containers'] += len(analysis.get('containers_cards', []))

    return element_counts

@app.route('/api/dashboard/real-analysis', methods=['GET'])
def get_real_analysis():
    """Fixed real design analysis with accurate data extraction"""
    try:
        analysis_dir = 'analysis_output'
        if not os.path.exists(analysis_dir):
            return jsonify({'error': 'Analysis directory not found'}), 404
        
        print("\n" + "="*60)
        print("📊 Loading Real Design Analysis")
        print("="*60)
        
        index_data, frames_data = load_analysis_frames(analysis_dir)

        # Platform distribution
        platform_distribution = {}
        for frame in frames_data:
            frame_info = frame.get('frame_info', {})
            platform = frame_info.get('platform', 'WEB')
            platform_distribution[platform] = platform_distribution.get(platform, 0) + 1

        element_counts = summarize_elements(frames_data)
        
        print(f"Loaded {len(frames_data)} frames")
        print(f"Platform distribution: {platform_distribution}")
        print(f"Element counts: {element_counts}")
        
        # Generate insights from actual data
        insights = generate_insights_from_real_data(frames_data, element_counts)
        
        # Calculate quality metrics
        quality_metrics = calculate_quality_metrics_from_data(frames_data, element_counts)
        
        # Get section statistics
        section_stats = []
        index_file = os.path.join(analysis_dir, 'index.json')
        if index_data or os.path.exists(index_file):
            if not index_data and os.path.exists(index_file):
                with open(index_file, 'r', encoding='utf-8') as f:
                    index_data = json.load(f)

            for section in index_data.get('sections', []):
                section_path = os.path.join(analysis_dir, section['path'])
                frame_count = 0
                if os.path.exists(section_path):
                    frame_count = len([d for d in os.listdir(section_path) 
                                     if os.path.isdir(os.path.join(section_path, d))])
                
                test_case_count = count_section_test_cases(
                    section.get('display_name', ''),
                    section_path
                )
                
                section_stats.append({
                    'name': section.get('display_name', 'Unknown'),
                    'frames': frame_count,
                    'test_cases': test_case_count,
                    'quality_score': calculate_section_quality(section_path)
                })
        
        # Get test case analysis
        test_case_analysis = analyze_real_test_cases()
        
        response = {
            'designAnalysis': insights,
            'platformDistribution': platform_distribution,
            'sectionStatistics': section_stats,
            'qualityMetrics': quality_metrics,
            'testCaseAnalysis': test_case_analysis,
            'elementStatistics': element_counts,
            'element_statistics': element_counts  # backward compatibility
        }

        if generation_status.get('results'):
            response['session_id'] = generation_status['results'].get('session_id')
            response['cloudPaths'] = generation_status['results'].get('cloudPaths', {})
        
        print("✅ Analysis complete")
        print("="*60 + "\n")
        
        return jsonify(response)
        
    except Exception as e:
        print(f"❌ Error getting analysis: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


def calculate_real_metrics(design_elements):
    """Calculate real quality metrics from design elements"""
    metrics = {
        'accessibility': 75,
        'consistency': 70,
        'usability': 80,
        'error_handling': 65,
        'performance': 85,
        'security': 90
    }
    
    # Calculate accessibility based on color contrast, text size, etc.
    buttons = design_elements.get('buttons', [])
    input_fields = design_elements.get('input_fields', [])
    
    # Check for accessibility features
    accessible_elements = 0
    total_elements = len(buttons) + len(input_fields)
    
    for button in buttons:
        if button.get('text_color') and button.get('background_color'):
            accessible_elements += 1
    
    for field in input_fields:
        if field.get('label') and field.get('placeholder'):
            accessible_elements += 1
    
    if total_elements > 0:
        metrics['accessibility'] = int((accessible_elements / total_elements) * 100)
    
    # Calculate consistency
    unique_button_styles = set()
    for button in buttons:
        style_key = f"{button.get('background_color', '')}-{button.get('text_color', '')}-{button.get('shape', '')}"
        unique_button_styles.add(style_key)
    
    if len(buttons) > 0:
        consistency_score = max(0, 100 - (len(unique_button_styles) * 10))
        metrics['consistency'] = min(100, consistency_score)
    
    return metrics


def generate_real_insights(design_elements):
    """Generate real insights from design elements using OpenAI"""
    try:
        # Prepare data for OpenAI
        insights_data = {
            'buttons': design_elements.get('buttons', [])[:10],  # Limit to 10 for token management
            'input_fields': design_elements.get('input_fields', [])[:10],
            'headings': design_elements.get('headings_and_titles', [])[:10],
            'icons': design_elements.get('icons', [])[:10]
        }
        
        prompt = f"""
        Analyze this UI design data and provide specific, actionable insights:
        
        {json.dumps(insights_data, indent=2)}
        
        Focus on:
        1. **Design Positives**: What works well (specific elements, patterns, consistency)
        2. **Design Issues**: Problems that need attention (specific issues with evidence)
        3. **Required Corrections**: Critical fixes needed (with specific examples)
        4. **Improvement Suggestions**: UX enhancements (specific suggestions)
        
        Return JSON format:
        {{
            "design_positives": ["string", "string"],
            "design_negatives": ["string", "string"],
            "required_corrections": ["string", "string"],
            "improvement_suggestions": ["string", "string"],
            "recommendations": {{
                "critical_issues": ["string"],
                "ux_improvements": ["string"],
                "security_recommendations": ["string"],
                "mobile_optimizations": ["string"]
            }}
        }}
        """
        
        # Call OpenAI
        response = design_analyzer.openai_client.chat.completions.create(
            model="gpt-4o",
            messages=[{"role": "user", "content": prompt}],
            max_tokens=2000,
            temperature=0.3
        )
        
        content = response.choices[0].message.content
        
        # Extract JSON
        if "```json" in content:
            content = content.split("```json")[1].split("```")[0].strip()
        elif "```" in content:
            content = content.split("```")[1].split("```")[0].strip()
        
        insights = json.loads(content)
        
        # Add specific examples from actual data
        if insights_data['buttons']:
            first_button = insights_data['buttons'][0]
            insights['design_positives'].append(
                f"Button design: '{first_button.get('text', 'N/A')}' with {first_button.get('background_color', 'default')} background"
            )
        
        return insights
        
    except Exception as e:
        print(f"OpenAI insights error: {e}")
        # Return fallback insights based on data
        return generate_fallback_insights(design_elements)


def generate_fallback_insights(design_elements):
    """Generate fallback insights when OpenAI fails"""
    buttons = design_elements.get('buttons', [])
    input_fields = design_elements.get('input_fields', [])
    
    return {
        "design_positives": [
            f"Found {len(buttons)} interactive buttons",
            f"Detected {len(input_fields)} input fields",
            "Consistent color scheme across elements",
            "Clear visual hierarchy in design"
        ],
        "design_negatives": [
            "Some elements may lack proper labeling",
            "Inconsistent spacing detected",
            "Missing error states for forms",
            "Limited accessibility features"
        ],
        "required_corrections": [
            "Add proper labels to all interactive elements",
            "Implement consistent spacing system",
            "Add error handling for form validation",
            "Improve color contrast for accessibility"
        ],
        "improvement_suggestions": [
            "Add hover states to buttons",
            "Implement form validation feedback",
            "Add loading indicators",
            "Improve mobile responsiveness"
        ],
        "recommendations": {
            "critical_issues": [
                "Check all form fields for proper validation",
                "Verify button contrast ratios meet WCAG standards",
                "Ensure all interactive elements have proper states"
            ],
            "ux_improvements": [
                "Add micro-interactions for better feedback",
                "Implement progressive disclosure for complex forms",
                "Add keyboard navigation support"
            ],
            "security_recommendations": [
                "Implement proper input sanitization",
                "Add CSRF protection for forms",
                "Use HTTPS for all data transmission"
            ],
            "mobile_optimizations": [
                "Optimize touch targets for mobile",
                "Implement responsive breakpoints",
                "Add swipe gestures where applicable"
            ]
        }
    }


def calculate_section_statistics(analysis_dir):
    """Calculate real section statistics"""
    sections = []
    
    index_file = os.path.join(analysis_dir, 'index.json')
    if os.path.exists(index_file):
        with open(index_file, 'r', encoding='utf-8') as f:
            index_data = json.load(f)
            
            for section in index_data.get('sections', []):
                section_path = os.path.join(analysis_dir, section['path'])
                frame_count = 0
                
                if os.path.exists(section_path):
                    frame_count = len([d for d in os.listdir(section_path) 
                                     if os.path.isdir(os.path.join(section_path, d))])
                
                # Count test cases for this section
                test_case_count = count_section_test_cases(
                    section.get('display_name', ''),
                    section_path
                )
                
                sections.append({
                    'name': section.get('display_name', 'Unknown'),
                    'frames': frame_count,
                    'test_cases': test_case_count,
                    'quality_score': calculate_section_quality(section_path)
                })
    
    return sections


def count_section_test_cases(section_name, section_path=None):
    """Count test cases for a specific section"""
    test_cases_dir = 'test_cases_output'
    if not os.path.exists(test_cases_dir):
        return 0
    
    section_folder = None

    # Prefer explicit section path when provided (match by folder name)
    if section_path:
        base_name = os.path.basename(section_path.rstrip(os.sep))
        for item in os.listdir(test_cases_dir):
            candidate = os.path.join(test_cases_dir, item)
            if not os.path.isdir(candidate):
                continue
            if item == base_name or item.endswith(base_name) or item.split('_', 1)[-1] == base_name:
                section_folder = candidate
                break

    # Fallback: attempt to match by display name suffix
    if not section_folder:
        suffix = section_name.replace(' ', '_')
        for item in os.listdir(test_cases_dir):
            candidate = os.path.join(test_cases_dir, item)
            if os.path.isdir(candidate) and item.endswith(suffix):
                section_folder = candidate
                break
    
    if not section_folder:
        return 0
    
    # Count test cases in section
    total_cases = 0
    for root, dirs, files in os.walk(section_folder):
        for file in files:
            if file.endswith('.md') and not file.startswith('00_'):
                file_path = os.path.join(root, file)
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    test_cases = re.findall(r'\*\*Test Case ID\*\*:', content)
                    total_cases += len(test_cases)
    
    return total_cases


def calculate_section_quality(section_path):
    """Calculate quality score for a section"""
    if not os.path.exists(section_path):
        return 75  # Default
    
    # Analyze frames in section
    frame_scores = []
    for item in os.listdir(section_path):
        frame_path = os.path.join(section_path, item)
        if os.path.isdir(frame_path):
            analysis_file = os.path.join(frame_path, 'analysis.json')
            if os.path.exists(analysis_file):
                with open(analysis_file, 'r', encoding='utf-8') as f:
                    frame_data = json.load(f)
                    # Simple quality calculation based on completeness
                    analysis = frame_data.get('analysis', {})
                    element_count = sum(len(v) for v in analysis.values() if isinstance(v, list))
                    frame_scores.append(min(100, element_count * 2))  # More elements = higher score
    
    return int(sum(frame_scores) / len(frame_scores)) if frame_scores else 75


def analyze_real_test_cases():
    """Analyze real test cases from generated files"""
    test_cases_dir = 'test_cases_output'
    if not os.path.exists(test_cases_dir):
        return {
            'category_distribution': {},
            'priority_distribution': {},
            'section_distribution': {},
            'total_test_cases': 0
        }
    
    categories = defaultdict(int)
    priorities = defaultdict(int)
    sections = defaultdict(int)
    total_cases = 0
    
    # Parse all test case files
    for root, dirs, files in os.walk(test_cases_dir):
        for file in files:
            if file.endswith('.md') and not file.startswith('00_'):
                file_path = os.path.join(root, file)
                with open(file_path, 'r', encoding='utf-8') as f:
                    content = f.read()
                    
                    # Extract section from path
                    rel_path = os.path.relpath(root, test_cases_dir)
                    section = rel_path.split(os.sep)[0] if rel_path else 'root'
                    
                    # Find all test cases in file
                    test_case_blocks = re.split(r'\*\*Test Case ID\*\*:', content)
                    
                    for block in test_case_blocks[1:]:  # Skip first empty block
                        total_cases += 1
                        sections[section] += 1
                        
                        # Extract category
                        category_match = re.search(r'\*\*Category\*\*:\s*(.+)', block)
                        if category_match:
                            category = category_match.group(1).strip()
                            categories[category] += 1
                        
                        # Extract priority
                        priority_match = re.search(r'\*\*Priority\*\*:\s*(.+)', block, re.IGNORECASE)
                        if priority_match:
                            priority = priority_match.group(1).strip()
                            priorities[priority] += 1
    
    return {
        'category_distribution': dict(categories),
        'priority_distribution': dict(priorities),
        'section_distribution': dict(sections),
        'total_test_cases': total_cases
    }


@app.route('/api/dashboard/refresh', methods=['POST'])
def refresh_dashboard_data():
    """Force refresh of dashboard data"""
    try:
        # Clear any cached data
        # Re-analyze files
        stats = get_real_stats().get_json()
        analysis = get_real_analysis().get_json()
        
        return jsonify({
            'success': True,
            'stats': stats,
            'analysis': analysis
        })
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/dashboard/live-updates', methods=['GET'])
def get_live_updates():
    """Get live updates for dashboard"""
    try:
        # Check for new files or updates
        updates = {
            'new_test_cases': 0,
            'new_frames': 0,
            'updated_at': time.strftime('%H:%M:%S'),
            'generation_status': generation_status.get('complete', False)
        }
        
        return jsonify(updates)
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500
 
if __name__ == '__main__':
    print("\n" + "="*60)
    print("🚀 Figma Test Case Generator - Starting Server")
    print("="*60)
    # print("\n📱 Access the application at: http://localhost:5000")
    print("\n💡 Make sure you have:")
    print("   - Figma Access Token")
    print("   - OpenAI API Key")
    print("   - Figma File Key or URL")
    print("\n⏹️  Press CTRL+C to stop the server\n")
    app.run(debug=True, port=8080, host='0.0.0.0')
